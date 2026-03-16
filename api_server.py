import base64
import io
import json
import os
import re
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload

    GOOGLE_DEPS_INSTALLED = True
except Exception:
    GOOGLE_DEPS_INSTALLED = False


PROJECT_ROOT = Path(__file__).resolve().parent
DB_NAME = os.getenv("TG_DB_NAME", "trade_guardian_v4.db")
DB_PATH = str(PROJECT_ROOT / DB_NAME)
SCOPES = ["https://www.googleapis.com/auth/drive"]
DRIVE_DB_FILE_NAME = os.getenv("DRIVE_DB_FILE_NAME", DB_NAME)
DRIVE_FOLDER_ID = os.getenv("DRIVE_FOLDER_ID", "").strip()

AUTO_EXCLUDE_GROUP_KEYWORDS = ("test", "testing", "experiment", "exp", "demo", "sandbox", "paper")

PARSER_COLUMN_ALIASES = {
    "name": {"name", "symbol", "trade_name"},
    "group": {"group", "strategy_group", "group_name"},
    "created_at": {"created_at", "created", "open_date", "entry_date", "date_opened", "opened_at"},
    "expiration": {"expiration", "expiry", "expires", "exp_date", "close_date", "closed_at", "exit_date"},
    "total_return": {"total_return", "total_return_dollar", "total_return_usd", "pnl", "p_l", "profit_loss", "net_pl"},
    "net_debit_credit": {"net_debit_credit", "net_debit", "debit_credit", "net_price", "debit", "credit"},
    "theta": {"theta"},
    "delta": {"delta"},
    "gamma": {"gamma"},
    "vega": {"vega"},
    "iv": {"iv", "implied_volatility"},
    "link": {"link", "url", "trade_url", "optionstrat_link"},
    "quantity": {"quantity", "qty", "contracts"},
    "entry_price": {"entry_price", "open_price", "entry", "price_open", "cost"},
    "close_price": {"close_price", "mark", "exit_price", "close", "price_close"},
}

DEFAULT_STRATEGIES = [
    ("130/160", "130/160", 500, 36, 0.8, "Income Discipline", 4000),
    ("160/190", "160/190", 700, 44, 0.8, "Patience Training", 5200),
    ("M200", "M200", 900, 41, 0.8, "Emotional Mastery", 8000),
    ("SMSF", "SMSF", 600, 40, 0.8, "Wealth Builder", 5000),
]

LAST_SYNC_RESULT: Dict = {}


class InMemoryUpload(io.BytesIO):
    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name or "upload"


def normalize_header_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def normalize_group_value(value) -> str:
    group = str(value).strip()
    if not group or group.lower() in {"nan", "none", "null"}:
        return "Uncategorized"
    return group


def clean_num(x):
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return 0.0
    s = re.sub(r"[,$%]", "", s)
    s = s.replace("(", "-").replace(")", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def detect_header_row_from_preview(preview_df: pd.DataFrame) -> int:
    if preview_df is None or preview_df.empty:
        return 0
    for idx, row in preview_df.head(35).iterrows():
        normalized_values = {normalize_header_key(v) for v in row.values if str(v).strip() != ""}
        if "name" in normalized_values and ("total_return" in normalized_values or "pnl" in normalized_values):
            return int(idx)
    return 0


def build_parser_column_map(columns) -> Dict:
    mapping = {}
    used_canonicals = set()
    for col in columns:
        norm = normalize_header_key(col)
        for canonical, aliases in PARSER_COLUMN_ALIASES.items():
            if norm in aliases and canonical not in used_canonicals:
                mapping[col] = canonical
                used_canonicals.add(canonical)
                break
    return mapping


def get_first_value(row: pd.Series, keys: List[str], default=None):
    for key in keys:
        if key in row and pd.notnull(row[key]) and str(row[key]).strip() != "":
            return row[key]
    return default


def get_db_connection():
    return sqlite3.connect(DB_PATH)


def add_column_safe(cursor, table: str, col_name: str, col_type: str):
    try:
        cursor.execute(f"SELECT {col_name} FROM {table} LIMIT 1")
    except Exception:
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass


def seed_default_strategies():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("SELECT count(*) FROM strategy_config")
        count = c.fetchone()[0]
        if count == 0:
            c.executemany("INSERT INTO strategy_config VALUES (?,?,?,?,?,?,?)", DEFAULT_STRATEGIES)
            conn.commit()
    finally:
        conn.close()


def init_db():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(
        """CREATE TABLE IF NOT EXISTS trades (
            id TEXT PRIMARY KEY,
            name TEXT,
            strategy TEXT,
            status TEXT,
            entry_date DATE,
            exit_date DATE,
            days_held INTEGER,
            debit REAL,
            lot_size INTEGER,
            pnl REAL,
            theta REAL,
            delta REAL,
            gamma REAL,
            vega REAL,
            notes TEXT,
            tags TEXT,
            parent_id TEXT,
            put_pnl REAL,
            call_pnl REAL,
            iv REAL,
            link TEXT,
            original_group TEXT
        )"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trade_id TEXT,
            snapshot_date DATE,
            pnl REAL,
            days_held INTEGER,
            theta REAL,
            delta REAL,
            vega REAL,
            gamma REAL,
            FOREIGN KEY(trade_id) REFERENCES trades(id)
        )"""
    )
    c.execute(
        """CREATE TABLE IF NOT EXISTS strategy_config (
            name TEXT PRIMARY KEY,
            identifier TEXT,
            target_pnl REAL,
            target_days INTEGER,
            min_stability REAL,
            description TEXT,
            typical_debit REAL
        )"""
    )
    add_column_safe(c, "snapshots", "theta", "REAL")
    add_column_safe(c, "snapshots", "delta", "REAL")
    add_column_safe(c, "snapshots", "vega", "REAL")
    add_column_safe(c, "snapshots", "gamma", "REAL")
    add_column_safe(c, "strategy_config", "typical_debit", "REAL")
    add_column_safe(c, "trades", "original_group", "TEXT")
    c.execute("CREATE INDEX IF NOT EXISTS idx_status ON trades(status)")
    conn.commit()
    conn.close()
    seed_default_strategies()


def load_strategy_config() -> Dict:
    init_db()
    conn = get_db_connection()
    try:
        df = pd.read_sql("SELECT * FROM strategy_config", conn)
        cfg = {}
        for _, row in df.iterrows():
            cfg[str(row["name"])] = {
                "id": str(row.get("identifier", "")),
                "pnl": clean_num(row.get("target_pnl", 0)),
                "dit": clean_num(row.get("target_days", 0)),
                "stability": clean_num(row.get("min_stability", 0)),
                "debit_per_lot": clean_num(row.get("typical_debit", 5000)) or 5000,
            }
        return cfg
    finally:
        conn.close()


def generate_id(name: str, strategy: str, entry_date) -> str:
    safe_name = re.sub(r"[^A-Z0-9]+", "_", str(name).upper())[:24]
    safe_strategy = re.sub(r"[^A-Z0-9]+", "_", str(strategy).upper())[:18]
    d_str = pd.to_datetime(entry_date).strftime("%Y%m%d")
    return f"{safe_name}_{safe_strategy}_{d_str}"


def get_strategy_dynamic(trade_name: str, group_name: str, config_dict: Dict) -> str:
    t_name = str(trade_name).upper().strip()
    g_name = str(group_name).upper().strip()
    sorted_strats = sorted(config_dict.items(), key=lambda x: len(str(x[1].get("id", ""))), reverse=True)
    for strat_name, details in sorted_strats:
        token = str(details.get("id", "")).upper().strip()
        if token and token in g_name:
            return strat_name
    for strat_name, details in sorted_strats:
        token = str(details.get("id", "")).upper().strip()
        if token and token in t_name:
            return strat_name
    return "Other"


def attach_excel_hyperlinks(file_obj: InMemoryUpload, header_row: int, df_raw: pd.DataFrame, column_map: Dict) -> pd.DataFrame:
    link_raw_col = None
    for raw_col, canonical in column_map.items():
        if canonical == "link":
            link_raw_col = raw_col
            break
    if link_raw_col is None:
        return df_raw

    try:
        file_obj.seek(0)
        wb = load_workbook(file_obj, data_only=False)
        sheet = wb.active
        excel_header_row = header_row + 1
        link_col_idx = None
        target_key = normalize_header_key(link_raw_col)
        for cell in sheet[excel_header_row]:
            cell_key = normalize_header_key(cell.value)
            if cell_key == target_key or cell_key in PARSER_COLUMN_ALIASES["link"]:
                link_col_idx = cell.col_idx
                break
        if not link_col_idx:
            return df_raw
        links = []
        for i in range(len(df_raw)):
            excel_row_idx = excel_header_row + 1 + i
            cell = sheet.cell(row=excel_row_idx, column=link_col_idx)
            url = ""
            if cell.hyperlink and cell.hyperlink.target:
                url = cell.hyperlink.target
            elif cell.value and str(cell.value).startswith("=HYPERLINK"):
                parts = str(cell.value).split('"')
                if len(parts) > 1:
                    url = parts[1]
            links.append(url if url else "")
        df_raw[link_raw_col] = links
    except Exception:
        pass
    return df_raw


def parse_optionstrat_file(file_obj: InMemoryUpload, file_type: str, config_dict: Dict) -> List[Dict]:
    try:
        file_name = str(getattr(file_obj, "name", "")).lower()
        df_raw = None
        header_row = 0
        if file_name.endswith((".xlsx", ".xls")):
            try:
                file_obj.seek(0)
                preview = pd.read_excel(file_obj, header=None, nrows=40)
                header_row = detect_header_row_from_preview(preview)
                file_obj.seek(0)
                df_raw = pd.read_excel(file_obj, header=header_row)
                column_map = build_parser_column_map(df_raw.columns)
                df_raw = attach_excel_hyperlinks(file_obj, header_row, df_raw, column_map)
            except Exception:
                df_raw = None

        if df_raw is None:
            try:
                file_obj.seek(0)
                preview_csv = pd.read_csv(file_obj, header=None, nrows=40, dtype=str, engine="python", on_bad_lines="skip")
                header_row = detect_header_row_from_preview(preview_csv)
            except Exception:
                header_row = 0
            file_obj.seek(0)
            df_raw = pd.read_csv(file_obj, skiprows=header_row)

        if df_raw is None or df_raw.empty:
            return []

        column_map = build_parser_column_map(df_raw.columns)
        df_norm = df_raw.rename(columns=column_map).copy()
        if "name" not in df_norm.columns or "total_return" not in df_norm.columns:
            return []

        def parse_leg_value(leg_row: pd.Series, column_name: str, fallback_idx: int) -> float:
            if column_name in leg_row and pd.notnull(leg_row[column_name]):
                return clean_num(leg_row[column_name])
            try:
                return clean_num(leg_row.iloc[fallback_idx])
            except Exception:
                return 0.0

        def finalize_trade(trade_data: pd.Series, legs: List[pd.Series], f_type: str) -> Optional[Dict]:
            if trade_data is None or trade_data.empty:
                return None
            name = str(trade_data.get("name", "")).strip()
            if not name or name.lower() in {"nan", "symbol"}:
                return None
            group = normalize_group_value(trade_data.get("group", ""))
            created = get_first_value(trade_data, ["created_at", "expiration"], "")
            try:
                start_dt = pd.to_datetime(created)
                if pd.isna(start_dt):
                    return None
            except Exception:
                return None

            strat = get_strategy_dynamic(name, group, config_dict)
            link = str(trade_data.get("link", "")).strip()
            if link.lower() in {"nan", "open"}:
                link = ""

            pnl = clean_num(trade_data.get("total_return", 0))
            debit = abs(clean_num(trade_data.get("net_debit_credit", 0)))
            theta = clean_num(trade_data.get("theta", 0))
            delta = clean_num(trade_data.get("delta", 0))
            gamma = clean_num(trade_data.get("gamma", 0))
            vega = clean_num(trade_data.get("vega", 0))
            iv = clean_num(trade_data.get("iv", 0))

            exit_dt = None
            raw_exp = trade_data.get("expiration")
            if pd.notnull(raw_exp) and str(raw_exp).strip() != "":
                try:
                    exit_dt = pd.to_datetime(raw_exp)
                except Exception:
                    pass

            if exit_dt is not None and f_type == "History":
                days_held = max(1, int((exit_dt - start_dt).days))
            else:
                days_held = max(1, int((datetime.now() - start_dt).days))

            strat_config = config_dict.get(strat, {})
            typical_debit = clean_num(strat_config.get("debit_per_lot", 5000)) or 5000
            lot_match = re.search(r"(\d+)\s*(?:LOT|L\b)", name, re.IGNORECASE)
            if lot_match:
                lot_size = int(lot_match.group(1))
            else:
                lot_size = int(round(debit / typical_debit)) if typical_debit else 1
            lot_size = max(1, lot_size)

            put_pnl = 0.0
            call_pnl = 0.0
            if f_type == "History":
                for leg in legs:
                    sym = str(leg.get("name", leg.iloc[0] if len(leg) > 0 else "")).strip()
                    if not sym.startswith("."):
                        continue
                    qty = parse_leg_value(leg, "quantity", 1)
                    entry = parse_leg_value(leg, "entry_price", 2)
                    close_price = parse_leg_value(leg, "close_price", 4)
                    leg_pnl = (close_price - entry) * qty * 100
                    if "P" in sym and "C" not in sym:
                        put_pnl += leg_pnl
                    elif "C" in sym and "P" not in sym:
                        call_pnl += leg_pnl
                    elif re.search(r"[0-9]P[0-9]", sym):
                        put_pnl += leg_pnl
                    elif re.search(r"[0-9]C[0-9]", sym):
                        call_pnl += leg_pnl

            return {
                "id": generate_id(name, strat, start_dt),
                "name": name,
                "strategy": strat,
                "start_dt": start_dt,
                "exit_dt": exit_dt,
                "days_held": days_held,
                "debit": debit,
                "lot_size": lot_size,
                "pnl": pnl,
                "theta": theta,
                "delta": delta,
                "gamma": gamma,
                "vega": vega,
                "iv": iv,
                "put_pnl": put_pnl,
                "call_pnl": call_pnl,
                "link": link,
                "group": group,
            }

        parsed_trades = []
        current_trade = None
        current_legs = []

        for _, row in df_norm.iterrows():
            name_val = str(row.get("name", "")).strip()
            if not name_val or name_val.lower() in {"nan", "symbol"}:
                continue
            if name_val.startswith("."):
                if current_trade is not None:
                    current_legs.append(row)
                continue
            if current_trade is not None:
                result = finalize_trade(current_trade, current_legs, file_type)
                if result:
                    parsed_trades.append(result)
            current_trade = row
            current_legs = []

        if current_trade is not None:
            result = finalize_trade(current_trade, current_legs, file_type)
            if result:
                parsed_trades.append(result)
        return parsed_trades
    except Exception:
        return []


def sync_data(file_obj: InMemoryUpload, file_type: str) -> Dict:
    conn = get_db_connection()
    c = conn.cursor()
    config_dict = load_strategy_config()
    logs = []
    totals = {"new_count": 0, "updated_count": 0, "snapshot_count": 0, "missing_marked": 0}

    db_active_ids = set()
    if file_type == "Active":
        try:
            current_active = pd.read_sql("SELECT id FROM trades WHERE status='Active'", conn)
            db_active_ids = set(current_active["id"].tolist())
        except Exception:
            pass

    file_found_ids = set()
    trades_data = parse_optionstrat_file(file_obj, file_type, config_dict)
    if not trades_data:
        logs.append(f"{file_obj.name}: Skipped (No valid trades found)")
        conn.close()
        return {"logs": logs, **totals}

    for t in trades_data:
        trade_id = t["id"]
        if file_type == "Active":
            file_found_ids.add(trade_id)

        c.execute(
            "SELECT id, status, theta, delta, gamma, vega, put_pnl, call_pnl, iv, link, lot_size, strategy FROM trades WHERE id = ?",
            (trade_id,),
        )
        existing = c.fetchone()

        if existing is None and t["link"] and len(t["link"]) > 15:
            c.execute("SELECT id, name FROM trades WHERE link = ?", (t["link"],))
            link_match = c.fetchone()
            if link_match:
                old_id, old_name = link_match
                try:
                    c.execute("UPDATE snapshots SET trade_id = ? WHERE trade_id = ?", (trade_id, old_id))
                    c.execute("UPDATE trades SET id=?, name=? WHERE id=?", (trade_id, t["name"], old_id))
                    logs.append(f"Renamed: '{old_name}' -> '{t['name']}'")
                    c.execute(
                        "SELECT id, status, theta, delta, gamma, vega, put_pnl, call_pnl, iv, link, lot_size, strategy FROM trades WHERE id = ?",
                        (trade_id,),
                    )
                    existing = c.fetchone()
                    if file_type == "Active":
                        if old_id in db_active_ids:
                            db_active_ids.remove(old_id)
                        db_active_ids.add(trade_id)
                except Exception:
                    pass

        status = "Active" if file_type == "Active" else "Expired"

        if existing is None:
            c.execute(
                """INSERT INTO trades
                (id, name, strategy, status, entry_date, exit_date, days_held, debit, lot_size, pnl, theta, delta, gamma, vega, notes, tags, parent_id, put_pnl, call_pnl, iv, link, original_group)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    trade_id,
                    t["name"],
                    t["strategy"],
                    status,
                    t["start_dt"].date(),
                    t["exit_dt"].date() if t["exit_dt"] is not None else None,
                    t["days_held"],
                    t["debit"],
                    t["lot_size"],
                    t["pnl"],
                    t["theta"],
                    t["delta"],
                    t["gamma"],
                    t["vega"],
                    "",
                    "",
                    "",
                    t["put_pnl"],
                    t["call_pnl"],
                    t["iv"],
                    t["link"],
                    t["group"],
                ),
            )
            totals["new_count"] += 1
        else:
            db_lot_size = existing[10]
            final_lot_size = db_lot_size if db_lot_size and db_lot_size > 0 else t["lot_size"]

            db_strategy = existing[11]
            final_strategy = t["strategy"] if db_strategy == "Other" and t["strategy"] != "Other" else db_strategy
            old_put = existing[6] if existing[6] else 0.0
            old_call = existing[7] if existing[7] else 0.0
            old_iv = existing[8] if existing[8] else 0.0
            old_link = existing[9] if existing[9] else ""
            old_status = existing[1]
            old_theta = existing[2]

            final_theta = t["theta"] if t["theta"] != 0 else old_theta
            final_delta = t["delta"] if t["delta"] != 0 else 0
            final_gamma = t["gamma"] if t["gamma"] != 0 else 0
            final_vega = t["vega"] if t["vega"] != 0 else 0
            final_iv = t["iv"] if t["iv"] != 0 else old_iv
            final_put = t["put_pnl"] if t["put_pnl"] != 0 else old_put
            final_call = t["call_pnl"] if t["call_pnl"] != 0 else old_call
            final_link = t["link"] if t["link"] else old_link

            if file_type == "History":
                c.execute(
                    """UPDATE trades SET
                    pnl=?, status=?, exit_date=?, days_held=?, theta=?, delta=?, gamma=?, vega=?, put_pnl=?, call_pnl=?, iv=?, link=?, lot_size=?, strategy=?, original_group=?
                    WHERE id=?""",
                    (
                        t["pnl"],
                        status,
                        t["exit_dt"].date() if t["exit_dt"] is not None else None,
                        t["days_held"],
                        final_theta,
                        final_delta,
                        final_gamma,
                        final_vega,
                        final_put,
                        final_call,
                        final_iv,
                        final_link,
                        final_lot_size,
                        final_strategy,
                        t["group"],
                        trade_id,
                    ),
                )
                totals["updated_count"] += 1
            elif old_status in ["Active", "Missing"]:
                c.execute(
                    """UPDATE trades SET
                    pnl=?, days_held=?, theta=?, delta=?, gamma=?, vega=?, iv=?, link=?, status='Active', exit_date=?, lot_size=?, strategy=?, original_group=?
                    WHERE id=?""",
                    (
                        t["pnl"],
                        t["days_held"],
                        final_theta,
                        final_delta,
                        final_gamma,
                        final_vega,
                        final_iv,
                        final_link,
                        t["exit_dt"].date() if t["exit_dt"] is not None else None,
                        final_lot_size,
                        final_strategy,
                        t["group"],
                        trade_id,
                    ),
                )
                totals["updated_count"] += 1

        if file_type == "Active":
            today = datetime.now().date()
            c.execute("SELECT id FROM snapshots WHERE trade_id=? AND snapshot_date=?", (trade_id, today))
            theta_val = t["theta"] if t["theta"] else 0.0
            delta_val = t["delta"] if t["delta"] else 0.0
            vega_val = t["vega"] if t["vega"] else 0.0
            gamma_val = t["gamma"] if t["gamma"] else 0.0
            if not c.fetchone():
                c.execute(
                    "INSERT INTO snapshots (trade_id, snapshot_date, pnl, days_held, theta, delta, vega, gamma) VALUES (?,?,?,?,?,?,?,?)",
                    (trade_id, today, t["pnl"], t["days_held"], theta_val, delta_val, vega_val, gamma_val),
                )
                totals["snapshot_count"] += 1
            else:
                c.execute(
                    "UPDATE snapshots SET pnl=?, days_held=?, theta=?, delta=?, vega=?, gamma=? WHERE trade_id=? AND snapshot_date=?",
                    (t["pnl"], t["days_held"], theta_val, delta_val, vega_val, gamma_val, trade_id, today),
                )

    if file_type == "Active" and file_found_ids:
        missing_ids = db_active_ids - file_found_ids
        if missing_ids:
            placeholders = ",".join("?" for _ in missing_ids)
            c.execute(f"UPDATE trades SET status = 'Missing' WHERE id IN ({placeholders})", list(missing_ids))
            totals["missing_marked"] = len(missing_ids)
            logs.append(f"Integrity: Marked {len(missing_ids)} trades as 'Missing'")

    conn.commit()
    conn.close()
    logs.append(f"{file_obj.name}: {totals['new_count']} New, {totals['updated_count']} Updated")
    return {"logs": logs, **totals}


class DriveManager:
    def __init__(self):
        self.creds = None
        self.service = None
        self.is_connected = False
        self.cached_file_id = os.getenv("DRIVE_FILE_ID", "").strip() or None
        self._init_connection()

    def _load_service_account_info(self) -> Optional[Dict]:
        raw_json = os.getenv("GCP_SERVICE_ACCOUNT_JSON", "").strip()
        if raw_json:
            return json.loads(raw_json)
        raw_b64 = os.getenv("GCP_SERVICE_ACCOUNT_B64", "").strip()
        if raw_b64:
            decoded = base64.b64decode(raw_b64).decode("utf-8")
            return json.loads(decoded)
        return None

    def _init_connection(self):
        if not GOOGLE_DEPS_INSTALLED:
            return
        try:
            service_info = self._load_service_account_info()
            if not service_info:
                return
            self.creds = service_account.Credentials.from_service_account_info(service_info, scopes=SCOPES)
            self.service = build("drive", "v3", credentials=self.creds, cache_discovery=False)
            self.is_connected = True
        except Exception:
            self.is_connected = False

    def _base_query(self) -> str:
        q = "trashed=false"
        if DRIVE_FOLDER_ID:
            q = f"'{DRIVE_FOLDER_ID}' in parents and {q}"
        return q

    def find_db_file(self) -> Tuple[Optional[str], Optional[str]]:
        if not self.is_connected:
            return None, None
        if self.cached_file_id:
            try:
                file = self.service.files().get(fileId=self.cached_file_id, fields="id,name").execute()
                return file["id"], file["name"]
            except Exception:
                self.cached_file_id = None
        try:
            query_exact = f"name='{DRIVE_DB_FILE_NAME}' and {self._base_query()}"
            results = self.service.files().list(q=query_exact, pageSize=1, fields="files(id, name)").execute()
            items = results.get("files", [])
            if items:
                self.cached_file_id = items[0]["id"]
                return items[0]["id"], items[0]["name"]
            query_fuzzy = f"name contains 'trade_guardian' and name contains '.db' and {self._base_query()}"
            results = self.service.files().list(q=query_fuzzy, pageSize=5, fields="files(id, name)").execute()
            items = results.get("files", [])
            if items:
                selected = items[0]
                self.cached_file_id = selected["id"]
                return selected["id"], selected["name"]
        except Exception:
            return None, None
        return None, None

    def get_cloud_modified_time(self, file_id: str) -> Optional[datetime]:
        try:
            file = self.service.files().get(fileId=file_id, fields="modifiedTime").execute()
            return datetime.strptime(file["modifiedTime"].replace("Z", "+0000"), "%Y-%m-%dT%H:%M:%S.%f%z")
        except Exception:
            return None

    def download_db(self, force=False) -> Tuple[bool, str]:
        if not self.is_connected:
            return False, "Drive not connected"
        file_id, file_name = self.find_db_file()
        if not file_id:
            return False, "No DB file found in Drive"
        if os.path.exists(DB_PATH) and not force:
            return True, "Local DB already exists"
        try:
            request = self.service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                _, done = downloader.next_chunk()
            with open(DB_PATH, "wb") as f:
                f.write(fh.getbuffer())
            return True, f"Downloaded '{file_name}'"
        except Exception as exc:
            return False, str(exc)

    def upload_db(self) -> Tuple[bool, str]:
        if not self.is_connected:
            return False, "Drive not connected"
        if not os.path.exists(DB_PATH):
            return False, "No local DB file to upload"
        file_id, file_name = self.find_db_file()
        media = MediaFileUpload(DB_PATH, mimetype="application/x-sqlite3", resumable=True)
        try:
            if file_id:
                self.service.files().update(fileId=file_id, media_body=media).execute()
                return True, f"Updated '{file_name}'"
            metadata = {"name": DRIVE_DB_FILE_NAME}
            if DRIVE_FOLDER_ID:
                metadata["parents"] = [DRIVE_FOLDER_ID]
            created = self.service.files().create(body=metadata, media_body=media, fields="id").execute()
            self.cached_file_id = created.get("id")
            return True, "Created new DB in Drive"
        except Exception as exc:
            return False, str(exc)


def get_db_version() -> Optional[str]:
    if not os.path.exists(DB_PATH):
        return None
    ts = os.path.getmtime(DB_PATH)
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


def get_db_meta() -> Dict:
    init_db()
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM trades")
        trade_count = int(c.fetchone()[0])
        c.execute("SELECT COUNT(*) FROM snapshots")
        snapshot_count = int(c.fetchone()[0])
        c.execute("SELECT MAX(snapshot_date) FROM snapshots")
        latest_snapshot = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM trades WHERE status='Active'")
        active_count = int(c.fetchone()[0])
        c.execute("SELECT COUNT(*) FROM trades WHERE status='Expired'")
        closed_count = int(c.fetchone()[0])
        c.execute("SELECT COUNT(*) FROM trades WHERE status='Missing'")
        missing_count = int(c.fetchone()[0])
        return {
            "trade_count": trade_count,
            "snapshot_count": snapshot_count,
            "latest_snapshot_date": latest_snapshot,
            "active_count": active_count,
            "closed_count": closed_count,
            "missing_count": missing_count,
            "db_version": get_db_version(),
        }
    finally:
        conn.close()


def build_app() -> Flask:
    app = Flask(__name__)
    cors_origins_raw = os.getenv("CORS_ORIGINS", "*")
    origins = [o.strip() for o in cors_origins_raw.split(",") if o.strip()]
    cors_value = "*" if origins == ["*"] or not origins else origins
    CORS(app, resources={r"/api/*": {"origins": cors_value}})
    drive_mgr = DriveManager()

    if not os.path.exists(DB_PATH):
        drive_mgr.download_db(force=True)
    init_db()

    @app.get("/")
    def serve_root():
        return send_from_directory(PROJECT_ROOT, "Trade_Guardian_Xlsx.html")

    @app.get("/xlsx")
    def serve_xlsx():
        return send_from_directory(PROJECT_ROOT, "Trade_Guardian_Xlsx.html")

    @app.get("/fixed")
    def serve_fixed():
        return send_from_directory(PROJECT_ROOT, "Trade_Guardian_Fixed.html")

    @app.get("/api/sync/status")
    def sync_status():
        return jsonify(
            {
                "ok": True,
                "service": "trade-guardian-api",
                "drive_connected": drive_mgr.is_connected,
                "drive_db_file_name": DRIVE_DB_FILE_NAME,
                "db_exists": os.path.exists(DB_PATH),
                "db_version": get_db_version(),
                "last_sync": LAST_SYNC_RESULT if LAST_SYNC_RESULT else None,
                "keywords_auto_exclude": list(AUTO_EXCLUDE_GROUP_KEYWORDS),
            }
        )

    @app.post("/api/sync/daily")
    def sync_daily():
        active_file = request.files.get("active_file")
        closed_file = request.files.get("closed_file")
        if active_file is None or closed_file is None:
            return jsonify({"ok": False, "error": "Both active_file and closed_file are required"}), 400

        init_db()
        active_data = active_file.read()
        closed_data = closed_file.read()
        active_upload = InMemoryUpload(active_data, active_file.filename or "active.xlsx")
        closed_upload = InMemoryUpload(closed_data, closed_file.filename or "closed.xlsx")

        active_result = sync_data(active_upload, "Active")
        closed_result = sync_data(closed_upload, "History")

        drive_ok = False
        drive_msg = "Drive not configured"
        if drive_mgr.is_connected:
            drive_ok, drive_msg = drive_mgr.upload_db()

        combined_logs = []
        combined_logs.extend(active_result.get("logs", []))
        combined_logs.extend(closed_result.get("logs", []))
        sync_time = datetime.now(timezone.utc).isoformat()
        meta = get_db_meta()
        result = {
            "ok": True,
            "sync_time": sync_time,
            "new_count": int(active_result["new_count"] + closed_result["new_count"]),
            "updated_count": int(active_result["updated_count"] + closed_result["updated_count"]),
            "missing_marked": int(active_result["missing_marked"]),
            "snapshot_count": int(active_result["snapshot_count"]),
            "drive_backup_status": {"ok": drive_ok, "message": drive_msg},
            "db_version": meta.get("db_version"),
            "logs": combined_logs,
            "meta": meta,
        }
        LAST_SYNC_RESULT.clear()
        LAST_SYNC_RESULT.update(result)
        return jsonify(result)

    @app.get("/api/db/download")
    def download_db():
        if not os.path.exists(DB_PATH):
            return jsonify({"ok": False, "error": "No DB file available"}), 404
        return send_file(DB_PATH, as_attachment=True, download_name=DRIVE_DB_FILE_NAME, mimetype="application/x-sqlite3")

    @app.get("/api/db/meta")
    def db_meta():
        return jsonify({"ok": True, **get_db_meta()})

    @app.get("/api/groups")
    def groups_meta():
        init_db()
        conn = get_db_connection()
        try:
            df = pd.read_sql("SELECT original_group FROM trades", conn)
            groups = sorted(
                {
                    normalize_group_value(v)
                    for v in df.get("original_group", pd.Series(dtype=str)).fillna("").astype(str).tolist()
                }
            )
            auto_excluded = [g for g in groups if any(k in g.lower() for k in AUTO_EXCLUDE_GROUP_KEYWORDS)]
            return jsonify({"ok": True, "groups": groups, "auto_excluded": auto_excluded})
        finally:
            conn.close()

    return app


app = build_app()


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=False)
