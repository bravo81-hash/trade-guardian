import streamlit as st
import pandas as pd
import numpy as np
import io
import plotly.express as px
import plotly.graph_objects as go
import sqlite3
import os
import re
import json
import time
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
from scipy import stats 
from scipy.spatial.distance import cdist 

# --- V150: ROBUST STATE INITIALIZATION ---
def initialize_session_state():
    """Improved State Management replacing globals()"""
    keys = {
        'velocity_stats': {},
        'mae_stats': {},
        'last_cloud_sync': None,
        'group_universe_excluded': [],
        'show_conflict': False,
        'show_pull_conflict': False
    }
    for key, default in keys.items():
        if key not in st.session_state:
            st.session_state[key] = default

initialize_session_state()
# -----------------------------------------

# --- GOOGLE DRIVE IMPORTS ---
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    GOOGLE_DEPS_INSTALLED = True
except ImportError:
    GOOGLE_DEPS_INSTALLED = False

# --- PAGE CONFIG ---
st.set_page_config(page_title="Trade Guardian XLSX", layout="wide", page_icon="🛡️")

# --- UI OVERHAUL: CSS INJECTION ---
def inject_custom_css():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

        :root {
            --tg-bg: #0f172a;
            --tg-bg-secondary: #111827;
            --tg-panel: rgba(30, 41, 59, 0.95);
            --tg-panel-strong: rgba(30, 41, 59, 0.98);
            --tg-border: rgba(51, 65, 85, 0.82);
            --tg-border-strong: rgba(51, 65, 85, 0.95);
            --tg-shadow: 0 20px 38px -28px rgba(0, 0, 0, 0.55);
            --tg-shadow-soft: 0 4px 6px -1px rgba(0, 0, 0, 0.2), 0 2px 4px -1px rgba(0, 0, 0, 0.1);
            --tg-text: #e2e8f0;
            --tg-muted: #94a3b8;
            --tg-accent: #6366f1;
            --tg-accent-soft: rgba(99, 102, 241, 0.14);
            --tg-accent-strong: #c7d2fe;
            --tg-positive: #10b981;
            --tg-negative: #f43f5e;
            --tg-warning: #f59e0b;
            --tg-font-body: clamp(14px, 0.35vw + 13px, 15.5px);
            --tg-font-title-xl: clamp(1.9rem, 2.6vw, 2.75rem);
            --tg-font-title-lg: clamp(1.4rem, 1.7vw, 2rem);
            --tg-font-title-md: clamp(1rem, 0.7vw, 1.12rem);
            --tg-font-label: 11px;
            --tg-leading-body: 1.52;
            --tg-leading-heading: 1.16;
            --tg-icon-stroke: 1.85;
            --tg-motion-fast: 180ms;
            --tg-motion-base: 240ms;
            --tg-motion-slow: 360ms;
            --tg-ease-standard: cubic-bezier(0.2, 0, 0, 1);
            --tg-ease-emphasized: cubic-bezier(0.2, 0, 0, 1);
        }

        * { box-sizing: border-box; }

        .stApp {
            font-family: 'Inter', sans-serif;
            background:
                radial-gradient(1200px 500px at -10% -20%, rgba(99, 102, 241, 0.12), transparent 55%),
                radial-gradient(900px 400px at 110% -10%, rgba(56, 189, 248, 0.10), transparent 60%),
                var(--tg-bg);
            color: var(--tg-text);
            text-rendering: optimizeLegibility;
            -webkit-font-smoothing: antialiased;
            -moz-osx-font-smoothing: grayscale;
        }

        [data-testid="stSidebar"] {
            background: rgba(15, 23, 42, 0.92);
            border-right: 1px solid var(--tg-border);
            backdrop-filter: blur(10px);
        }

        h1, h2, h3, h4, h5, h6 {
            color: var(--tg-text) !important;
            letter-spacing: -0.025em;
            line-height: var(--tg-leading-heading);
            font-weight: 700 !important;
        }

        p, label, span, div, li {
            font-family: 'Inter', sans-serif;
        }

        code, pre, .font-mono {
            font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
        }

        .gradient-text {
            background: linear-gradient(to right, #6366f1, #38bdf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            font-weight: 800;
        }

        button, a, input, select, textarea {
            transition:
                border-color var(--tg-motion-fast) var(--tg-ease-standard),
                background-color var(--tg-motion-fast) var(--tg-ease-standard),
                color var(--tg-motion-fast) var(--tg-ease-standard),
                transform var(--tg-motion-base) var(--tg-ease-standard),
                box-shadow var(--tg-motion-base) var(--tg-ease-standard);
        }

        button:focus-visible, a:focus-visible, input:focus-visible, select:focus-visible, textarea:focus-visible {
            outline: 2px solid var(--tg-accent);
            outline-offset: 2px;
        }

        div[data-testid="stMetric"] {
            background: var(--tg-panel);
            border: 1px solid var(--tg-border);
            border-radius: 1rem;
            box-shadow: var(--tg-shadow-soft);
            backdrop-filter: blur(10px);
            padding: 0.95rem 1rem !important;
            min-height: 84px;
        }

        div[data-testid="stMetric"]:hover {
            transform: translateY(-1px);
            border-color: rgba(99, 102, 241, 0.28);
            box-shadow: 0 10px 30px -16px rgba(79, 70, 229, 0.24);
        }

        [data-testid="stMetricLabel"] {
            color: var(--tg-muted) !important;
            font-size: 0.75rem !important;
            margin-bottom: 0 !important;
        }

        [data-testid="stMetricValue"] {
            color: var(--tg-text) !important;
            font-size: 1.55rem !important;
            font-weight: 800 !important;
            line-height: 1.1;
        }

        [data-testid="stMetricDelta"] {
            font-size: 0.78rem !important;
        }

        div[data-testid="stExpander"] {
            background: var(--tg-panel);
            border: 1px solid var(--tg-border);
            border-radius: 1rem;
            box-shadow: var(--tg-shadow-soft);
        }

        .stTabs [data-baseweb="tab-list"] {
            gap: 0.5rem;
        }

        .stTabs [data-baseweb="tab"] {
            height: 46px;
            white-space: pre-wrap;
            background: transparent;
            border-radius: 0.9rem;
            color: var(--tg-muted);
            font-weight: 700 !important;
            font-size: 0.94rem;
            text-transform: uppercase;
            letter-spacing: 0.08em;
        }

        .stTabs [data-baseweb="tab"]:hover {
            color: var(--tg-accent-strong);
            background: rgba(99, 102, 241, 0.12);
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(99, 102, 241, 0.24), rgba(99, 102, 241, 0.14)) !important;
            color: var(--tg-accent-strong) !important;
            border-bottom: 1px solid rgba(129, 140, 248, 0.45);
            box-shadow: inset 0 -1px 0 rgba(129, 140, 248, 0.45);
        }

        .stButton > button,
        .stDownloadButton > button,
        .stFileUploader button,
        [data-testid="stSidebar"] button {
            border-radius: 0.9rem !important;
            font-weight: 700 !important;
            border: 1px solid transparent;
            box-shadow: none;
        }

        .stButton > button {
            background: var(--tg-accent);
            color: #fff;
        }

        .stButton > button:hover {
            background: var(--tg-accent-strong);
            box-shadow: 0 14px 24px -20px rgba(99, 102, 241, 0.45);
        }

        .stDownloadButton > button {
            background: rgba(16, 185, 129, 0.14);
            color: #a7f3d0;
            border-color: rgba(16, 185, 129, 0.24);
        }

        .stDownloadButton > button:hover {
            background: rgba(16, 185, 129, 0.2);
        }

        [data-testid="stFileUploader"] section {
            border-radius: 1rem;
            border: 1px dashed rgba(124, 147, 175, 0.42);
            background: linear-gradient(180deg, rgba(30, 41, 59, 0.9), rgba(15, 23, 42, 0.8));
            padding: 0.4rem;
        }

        [data-testid="stFileUploader"] small {
            color: var(--tg-muted);
        }

        [data-baseweb="input"] > div,
        [data-baseweb="textarea"] textarea,
        [data-baseweb="select"] > div {
            border-radius: 0.9rem !important;
        }

        [data-testid="stDataFrame"] {
            border: 1px solid var(--tg-border);
            border-radius: 1rem;
            overflow: hidden;
            box-shadow: var(--tg-shadow-soft);
        }

        [data-baseweb="notification"] {
            border-radius: 0.9rem;
        }

        hr {
            border-color: var(--tg-border);
            opacity: 0.85;
        }

        .stApp [data-testid="stVerticalBlock"] {
            gap: 0.9rem;
        }

        .pro-card {
            border-radius: 1.25rem;
            border: 1px solid rgba(51, 65, 85, 0.82);
            background: linear-gradient(180deg, rgba(30, 41, 59, 0.98), rgba(15, 23, 42, 0.94));
            box-shadow: 0 16px 36px -28px rgba(0, 0, 0, 0.45);
        }

        .micro-pill {
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
            padding: 0.25rem 0.6rem;
            border-radius: 9999px;
            font-size: 11px;
            font-weight: 700;
            letter-spacing: 0.02em;
            border: 1px solid rgba(129, 140, 248, 0.28);
            background: rgba(99, 102, 241, 0.14);
            color: #c7d2fe;
        }

        .gradient-text {
            background: linear-gradient(to right, #c7d2fe, #38bdf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            font-weight: 800;
        }

        ::-webkit-scrollbar { width: 8px; height: 8px; }
        ::-webkit-scrollbar-track { background: transparent; }
        ::-webkit-scrollbar-thumb { background: rgba(126, 145, 169, 0.62); border-radius: 9999px; }
        ::-webkit-scrollbar-thumb:hover { background: rgba(97, 122, 154, 0.82); }

        @media (max-width: 768px) {
            [data-testid="stMetricValue"] { font-size: 1.35rem !important; }
            .stTabs [data-baseweb="tab"] { font-size: 0.8rem; height: 42px; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

inject_custom_css()


def apply_chart_theme(fig):
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family="Inter, sans-serif", color="#475569"),
        title_font=dict(family="Inter, sans-serif", size=20, color="#0f172a"),
        hoverlabel=dict(bgcolor="#1e293b", font_size=13, font_family="Inter"),
        colorway=['#6366f1', '#10b981', '#38bdf8', '#f59e0b', '#f43f5e'],
        xaxis=dict(showgrid=True, gridcolor='rgba(148, 163, 184, 0.16)', zerolinecolor='rgba(148, 163, 184, 0.24)'),
        yaxis=dict(showgrid=True, gridcolor='rgba(148, 163, 184, 0.16)', zerolinecolor='rgba(148, 163, 184, 0.24)'),
    )
    return fig

# --- DEBUG BANNER ---
st.markdown("""
    <div style="background: rgba(99, 102, 241, 0.14); border: 1px solid rgba(129, 140, 248, 0.26); border-radius: 9999px; padding: 0.5rem 0.9rem; margin-bottom: 1rem; font-size: 0.78rem; color: #c7d2fe; display: inline-flex; align-items: center; gap: 0.55rem; font-weight: 700;">
        <span style="width:0.55rem;height:0.55rem;border-radius:9999px;background:#10b981;display:inline-block;"></span>
        <span>RUNNING VERSION: v151.0 (Strategy-first cockpit + cleaned signal stack)</span>
    </div>
""", unsafe_allow_html=True)

# --- HERO HEADER ---
st.markdown("""
    <div class="pro-card" style="margin-bottom: 1.35rem; padding: 1.35rem 1.45rem;">
        <div class="micro-pill" style="margin-bottom: 0.8rem;">Workbook Intelligence Mode</div>
        <h1 style="font-size: var(--tg-font-title-xl); line-height: 1.08; margin-bottom: 0.5rem;">
            Allantis <span class="gradient-text">Trade Guardian</span>
        </h1>
        <p style="max-width: 56rem; color: var(--tg-muted); margin: 0;">
            Streamlined two-workbook workflow for all active and all closed OptionStrat exports, with group universe controls and the same cleaner Inter-based visual language as the XLSX viewer.
        </p>
    </div>
""", unsafe_allow_html=True)

# --- DATABASE CONSTANTS ---
DB_NAME = "trade_guardian_v4.db"
SCOPES = ['https://www.googleapis.com/auth/drive']

# --- V150: CLOUD SYNC ENGINE w/ EXPONENTIAL BACKOFF ---
class DriveManager:
    def __init__(self):
        self.creds = None
        self.service = None
        self.is_connected = False
        self.cached_file_id = None
        
        if 'gcp_service_account' in st.secrets:
            try:
                service_account_info = st.secrets["gcp_service_account"]
                self.creds = service_account.Credentials.from_service_account_info(
                    service_account_info, scopes=SCOPES)
                self.service = build('drive', 'v3', credentials=self.creds)
                self.is_connected = True
            except Exception as e:
                st.error(f"Cloud Config Error: {e}")

    def list_files_debug(self):
        if not self.is_connected: return []
        try:
            results = self.service.files().list(pageSize=10, fields="files(id, name, modifiedTime)").execute()
            return results.get('files', [])
        except Exception as e:
            return []

    def find_db_file(self):
        if not self.is_connected: return None, None
        if self.cached_file_id:
            try:
                file = self.service.files().get(fileId=self.cached_file_id, fields='id,name').execute()
                return file['id'], file['name']
            except:
                self.cached_file_id = None

        try:
            query_exact = f"name='{DB_NAME}' and trashed=false"
            results = self.service.files().list(q=query_exact, pageSize=1, fields="files(id, name)").execute()
            items = results.get('files', [])
            if items: 
                self.cached_file_id = items[0]['id']
                return items[0]['id'], items[0]['name']

            query_fuzzy = "name contains 'trade_guardian' and name contains '.db' and trashed=false"
            results = self.service.files().list(q=query_fuzzy, pageSize=5, fields="files(id, name)").execute()
            items = results.get('files', [])
            
            if items:
                selected = items[0]
                for item in items:
                    if item['name'].startswith("trade_guardian_v4"):
                        selected = item
                        break
                self.cached_file_id = selected['id']
                return selected['id'], selected['name']
            return None, None
        except Exception as e:
            st.error(f"Drive Search Error: {e}")
            return None, None

    def get_cloud_modified_time(self, file_id):
        try:
            file = self.service.files().get(fileId=file_id, fields='modifiedTime').execute()
            dt = datetime.strptime(file['modifiedTime'].replace('Z', '+0000'), '%Y-%m-%dT%H:%M:%S.%f%z')
            return dt
        except:
            return None

    def create_backup(self, file_id, file_name):
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"BACKUP_{timestamp}_{file_name}"
            orig_file = self.service.files().get(fileId=file_id, fields='parents').execute()
            parents = orig_file.get('parents', [])
            metadata = {'name': backup_name}
            if parents: metadata['parents'] = parents
            self.service.files().copy(fileId=file_id, body=metadata).execute()
            self.cleanup_backups(file_name)
            return True
        except Exception as e:
            print(f"Backup failed: {e}")
            return False

    def cleanup_backups(self, original_name):
        try:
            query = f"name contains 'BACKUP_' and name contains '{original_name}' and trashed=false"
            results = self.service.files().list(q=query, pageSize=20, fields="files(id, name, createdTime)", orderBy="createdTime desc").execute()
            items = results.get('files', [])
            if len(items) > 5:
                for item in items[5:]:
                    try:
                        self.service.files().delete(fileId=item['id']).execute()
                    except: pass
        except: pass

    def download_db(self, force=False):
        file_id, file_name = self.find_db_file()
        if not file_id:
            return False, "Database not found in Cloud."
        
        if os.path.exists(DB_NAME) and not force:
            try:
                local_ts = os.path.getmtime(DB_NAME)
                local_mod = datetime.fromtimestamp(local_ts, tz=timezone.utc)
                cloud_time = self.get_cloud_modified_time(file_id)
                if cloud_time and (local_mod > cloud_time + timedelta(minutes=2)):
                    return False, f"CONFLICT: Your local database is NEWER ({local_mod.strftime('%H:%M')}) than the cloud file ({cloud_time.strftime('%H:%M')}).\n\nIf you pull now, you will lose your recent local changes."
            except Exception as e:
                print(f"Pull check warning: {e}")

        try:
            try: sqlite3.connect(DB_NAME).close()
            except: pass

            request = self.service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            
            with open(DB_NAME, "wb") as f:
                f.write(fh.getbuffer())
            
            st.session_state['last_cloud_sync'] = datetime.now()
            return True, f"Downloaded '{file_name}' successfully."
        except Exception as e:
            return False, str(e)

    def upload_db(self, force=False, retries=3):
        if not os.path.exists(DB_NAME):
            return False, "No local database found to upload."
            
        file_id, file_name = self.find_db_file()
        
        if file_id and not force:
            cloud_time = self.get_cloud_modified_time(file_id)
            local_ts = os.path.getmtime(DB_NAME)
            local_time = datetime.fromtimestamp(local_ts, tz=timezone.utc) 
            if cloud_time and (cloud_time > local_time + timedelta(seconds=2)):
                 return False, f"CONFLICT: Cloud file is newer ({cloud_time.strftime('%H:%M')}) than local ({local_time.strftime('%H:%M')}). Please Pull first."

        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute("PRAGMA integrity_check")
            result = cursor.fetchone()
            conn.close()
            if result[0] != "ok":
                return False, "❌ Local Database Corrupt. Do not upload."
        except Exception as e:
            return False, f"❌ DB Check Failed: {e}"

        media = MediaFileUpload(DB_NAME, mimetype='application/x-sqlite3', resumable=True)
        
        # --- V150: EXPONENTIAL BACKOFF LOGIC ---
        delay = 1
        for attempt in range(retries + 1):
            try:
                if file_id:
                    if attempt == 0: 
                        self.create_backup(file_id, file_name)
                    self.service.files().update(fileId=file_id, media_body=media).execute()
                    action = f"Updated '{file_name}' (Backup created)"
                else:
                    file_metadata = {'name': DB_NAME}
                    self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                    action = "Created New File"
                
                st.session_state['last_cloud_sync'] = datetime.now()
                return True, f"Sync Successful: {action}"
            except Exception as e:
                if attempt < retries:
                    time.sleep(delay)
                    delay *= 2  # Double the delay each time
                    continue
                return False, f"Upload failed after {retries} retries: {str(e)}"

# Initialize Drive Manager
drive_mgr = DriveManager()

def auto_sync_if_connected():
    if not drive_mgr.is_connected: return
    with st.spinner("☁️ Auto-syncing to cloud..."):
        success, msg = drive_mgr.upload_db()
        if success:
            st.toast(f"✅ Cloud Saved: {datetime.now().strftime('%H:%M')}")
        elif "CONFLICT" in msg:
            st.error(f"⚠️ Auto-sync BLOCKED: Conflict detected. Please resolve in sidebar.")
        else:
            st.warning(f"⚠️ Auto-sync failed: {msg}")

# --- DATABASE ENGINE ---
def get_db_connection():
    return sqlite3.connect(DB_NAME)

def init_db():
    if not os.path.exists(DB_NAME) and drive_mgr.is_connected:
        success, msg = drive_mgr.download_db()
        if success:
            st.toast(f"☁️ Cloud Data Loaded: {msg}")
    
    conn = get_db_connection()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS trades (
                    id TEXT PRIMARY KEY,
                    name TEXT,
                    strategy TEXT,
                    status TEXT,
                    entry_date DATE,
                    exit_date DATE,
                    days_held INTEGER,
                    debit REAL,
                    lot_size INTEGER,
                    pnl REAL,
                    theta REAL,
                    delta REAL,
                    gamma REAL,
                    vega REAL,
                    notes TEXT,
                    tags TEXT,
                    parent_id TEXT,
                    put_pnl REAL,
                    call_pnl REAL,
                    iv REAL,
                    link TEXT,
                    original_group TEXT
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS snapshots (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    trade_id TEXT,
                    snapshot_date DATE,
                    pnl REAL,
                    days_held INTEGER,
                    theta REAL,
                    delta REAL,
                    vega REAL,
                    gamma REAL, 
                    FOREIGN KEY(trade_id) REFERENCES trades(id)
                )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS strategy_config (
                    name TEXT PRIMARY KEY,
                    identifier TEXT,
                    target_pnl REAL,
                    target_days INTEGER,
                    min_stability REAL,
                    description TEXT,
                    typical_debit REAL
                )''')
    
    def add_column_safe(table, col_name, col_type):
        try:
            c.execute(f"SELECT {col_name} FROM {table} LIMIT 1")
        except:
            try:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
            except: pass

    add_column_safe('snapshots', 'theta', 'REAL')
    add_column_safe('snapshots', 'delta', 'REAL')
    add_column_safe('snapshots', 'vega', 'REAL')
    add_column_safe('snapshots', 'gamma', 'REAL')
    add_column_safe('strategy_config', 'typical_debit', 'REAL')
    add_column_safe('trades', 'original_group', 'TEXT')
    
    c.execute("CREATE INDEX IF NOT EXISTS idx_status ON trades(status)")
    conn.commit()
    conn.close()
    seed_default_strategies()

def seed_default_strategies(force_reset=False):
    conn = get_db_connection()
    c = conn.cursor()
    try:
        if force_reset:
            c.execute("DELETE FROM strategy_config")
        
        c.execute("SELECT count(*) FROM strategy_config")
        count = c.fetchone()[0]
        
        if count == 0:
            defaults = [
                ('130/160', '130/160', 500, 36, 0.8, 'Income Discipline', 4000),
                ('160/190', '160/190', 700, 44, 0.8, 'Patience Training', 5200),
                ('M200', 'M200', 900, 41, 0.8, 'Emotional Mastery', 8000),
                ('SMSF', 'SMSF', 600, 40, 0.8, 'Wealth Builder', 5000)
            ]
            c.executemany("INSERT INTO strategy_config VALUES (?,?,?,?,?,?,?)", defaults)
            conn.commit()
            if force_reset:
                st.toast("Strategies Reset to Factory Defaults.")
    except Exception as e:
        print(f"Seeding error: {e}")
    finally:
        conn.close()

@st.cache_data(ttl=60)
def load_strategy_config():
    if not os.path.exists(DB_NAME): return {}
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='strategy_config'")
        if not c.fetchone(): return {}

        df = pd.read_sql("SELECT * FROM strategy_config", conn)
        config = {}
        for _, row in df.iterrows():
            typ_debit = row['typical_debit'] if 'typical_debit' in row and pd.notnull(row['typical_debit']) else 5000
            
            config[row['name']] = {
                'id': row['identifier'],
                'pnl': row['target_pnl'],
                'dit': row['target_days'],
                'stability': row['min_stability'],
                'debit_per_lot': typ_debit
            }
        return config
    except: return {}
    finally: conn.close()

# --- V150 QUANT LOGIC & INTELLIGENCE FUNCTIONS ---

def calculate_kelly_fraction(win_rate, avg_win, avg_loss, correlation_penalty=0.5):
    """
    V150: Modified Kelly Criterion for Options Portfolios.
    Applies a correlation penalty (Half-Kelly equivalent) to prevent massive over-leveraging.
    """
    if avg_loss == 0 or avg_win <= 0:
        return 0.0
    b = abs(avg_win / avg_loss)
    kelly = (win_rate * b - (1 - win_rate)) / b
    # Cap at 25% generally, scale by correlation_penalty to reflect options dependence
    return max(0, min(kelly * correlation_penalty, 0.25))

def vectorized_theta_efficiency(df):
    """
    V150: Vectorized efficiency logic, safely avoiding division by zero.
    """
    if df.empty: return df
    theta_safe = np.where(df['Theta'] == 0, 1, df['Theta'])
    days_safe = np.where(df['Days Held'] == 0, 1, df['Days Held'])
    df['Theta_Efficiency_Score'] = df['P&L'] / (theta_safe * days_safe)
    return df

def calculate_portfolio_var(df, capital, confidence=0.95):
    """
    V151: Historical realized 1-day VaR in dollars.
    Uses daily realized PnL (grouped by Exit Date), converts to return on provided capital,
    and returns a negative dollar amount.
    """
    if df.empty or capital <= 0:
        return 0.0
    if 'Exit Date' not in df.columns or 'P&L' not in df.columns:
        return 0.0

    daily = df.dropna(subset=['Exit Date']).copy()
    if daily.empty:
        return 0.0

    daily['Exit Date'] = pd.to_datetime(daily['Exit Date'], errors='coerce')
    daily = daily.dropna(subset=['Exit Date'])
    if daily.empty:
        return 0.0

    pnl_by_day = daily.groupby(daily['Exit Date'].dt.date)['P&L'].sum().sort_index()
    if len(pnl_by_day) < 2:
        return 0.0

    daily_returns = pnl_by_day / float(capital)
    mu = float(daily_returns.mean())
    sigma = float(daily_returns.std(ddof=0))
    if sigma == 0:
        return 0.0

    var_return = stats.norm.ppf(1 - confidence, mu, sigma)
    var_dollars = var_return * float(capital)
    return min(0.0, float(var_dollars))

def generate_trade_predictions(active_df, history_df, prob_low, prob_high, total_capital=100000):
    if active_df.empty or history_df.empty: return pd.DataFrame()
    features = ['Theta/Cap %', 'Delta', 'Debit/Lot']
    train_df = history_df.dropna(subset=features).copy()
    if len(train_df) < 5: return pd.DataFrame()
    predictions = []
    for _, row in active_df.iterrows():
        curr_vec = np.nan_to_num(row[features].values.astype(float)).reshape(1, -1)
        hist_vecs = np.nan_to_num(train_df[features].values.astype(float))
        distances = cdist(curr_vec, hist_vecs, metric='euclidean')[0]
        top_k_idx = np.argsort(distances)[:7]
        nearest_neighbors = train_df.iloc[top_k_idx]
        win_prob = (nearest_neighbors['P&L'] > 0).mean()
        avg_pnl = nearest_neighbors['P&L'].mean()
        avg_win = nearest_neighbors[nearest_neighbors['P&L'] > 0]['P&L'].mean()
        avg_loss = nearest_neighbors[nearest_neighbors['P&L'] < 0]['P&L'].mean()
        if pd.isna(avg_win): avg_win = 0
        if pd.isna(avg_loss): avg_loss = -avg_pnl * 0.5
        
        kelly_size = calculate_kelly_fraction(win_prob, avg_win, avg_loss)
        rec_dollars = kelly_size * total_capital
        avg_dist = distances[top_k_idx].mean()
        confidence = max(0, 100 - (avg_dist * 10)) 
        rec = "HOLD"
        if win_prob * 100 < prob_low: rec = "REDUCE/CLOSE"
        elif win_prob * 100 > prob_high: rec = "PRESS WINNER"
        predictions.append({
            'Trade Name': row['Name'], 'Strategy': row['Strategy'], 'Win Prob %': win_prob * 100,
            'Expected PnL': avg_pnl, 'Kelly Size %': kelly_size * 100, 'Rec. Size ($)': rec_dollars,
            'AI Rec': rec, 'Confidence': confidence
        })
    return pd.DataFrame(predictions)

def get_mae_stats(conn):
    query = """
    SELECT t.strategy, MIN(s.pnl) as worst_drawdown
    FROM snapshots s
    JOIN trades t ON s.trade_id = t.id
    WHERE t.status = 'Expired' AND t.pnl > 0
    GROUP BY t.id, t.strategy
    """
    try:
        df = pd.read_sql(query, conn)
        mae_stats = {}
        if not df.empty:
            for strategy in df['strategy'].unique():
                strat_df = df[df['strategy'] == strategy]
                if not strat_df.empty:
                    limit = strat_df['worst_drawdown'].quantile(0.05) 
                    mae_stats[strategy] = limit
        return mae_stats
    except Exception as e:
        return {}

def get_velocity_stats(expired_df):
    velocity_stats = {}
    if expired_df.empty: return velocity_stats
    winners = expired_df[expired_df['P&L'] > 0].copy()
    if winners.empty: return velocity_stats
    winners['days_held'] = winners['days_held'].replace(0, 1)
    winners['velocity'] = winners['P&L'] / winners['days_held']
    for strategy in winners['Strategy'].unique():
        s_df = winners[winners['Strategy'] == strategy]
        if len(s_df) > 2:
            mean_v = s_df['velocity'].mean()
            std_v = s_df['velocity'].std()
            velocity_stats[strategy] = {
                'threshold': mean_v + (2 * std_v),
                'mean': mean_v
            }
    return velocity_stats

def check_rot_and_efficiency(row, hist_avg_days):
    try:
        current_pnl = row['P&L']
        days = row.get('Days', 1)
        if days == 0: days = 1
        theta = row.get('Net Theta', 0)
        current_speed = current_pnl / days
        
        if theta != 0: theta_efficiency = (current_speed / theta) 
        else: theta_efficiency = 0
            
        status = "Healthy"
        if days > hist_avg_days * 1.2 and current_pnl < 0:
            status = "Rotting (Time > Avg & Red)"
        elif theta_efficiency < 0.2 and days > 10 and current_pnl > 0:
             status = "Inefficient (Theta Stuck)"
        elif theta_efficiency < 0 and days > 5:
             status = "Bleeding (Negative Efficiency)"
             
        return current_speed, theta_efficiency, status
    except:
        return 0, 0, "Error"

# --- HELPER FUNCTIONS ---
def get_strategy_dynamic(trade_name, group_name, config_dict):
    t_name = str(trade_name).upper().strip()
    g_name = str(group_name).upper().strip()
    sorted_strats = sorted(config_dict.items(), key=lambda x: len(str(x[1]['id'])), reverse=True)
    for strat_name, details in sorted_strats:
        key = str(details['id']).upper()
        if key in t_name: return strat_name
    for strat_name, details in sorted_strats:
        key = str(details['id']).upper()
        if key in g_name: return strat_name
    return "Other"

def clean_num(x):
    try:
        if pd.isna(x) or str(x).strip() == "": return 0.0
        val_str = str(x).replace('$', '').replace(',', '').replace('%', '').strip()
        val = float(val_str)
        if np.isnan(val): return 0.0
        return val
    except: return 0.0

def safe_fmt(val, fmt_str):
    try:
        if isinstance(val, (int, float)): return fmt_str.format(val)
        return str(val)
    except: return str(val)

def generate_id(name, strategy, entry_date):
    d_str = pd.to_datetime(entry_date).strftime('%Y%m%d')
    safe_name = re.sub(r'\W+', '', str(name))
    return f"{safe_name}_{strategy}_{d_str}"

def extract_ticker(name):
    try:
        parts = str(name).split(' ')
        if parts:
            ticker = parts[0].replace('.', '').upper()
            if ticker in ['M200', '130', '160', 'IRON', 'VERTICAL', 'SMSF']:
                return "UNKNOWN"
            return ticker
        return "UNKNOWN"
    except: return "UNKNOWN"

def theta_decay_model(initial_theta, days_held, strategy, dte_at_entry=45):
    t_frac = min(1.0, days_held / dte_at_entry) if dte_at_entry > 0 else 1.0
    if strategy in ['M200', '130/160', '160/190', 'SMSF']:
        if t_frac < 0.5: decay_factor = 1 - (2 * t_frac) ** 2
        else: decay_factor = 2 * (1 - t_frac)
        return initial_theta * max(0, decay_factor)
    elif 'VERTICAL' in str(strategy).upper() or 'DIRECTIONAL' in str(strategy).upper():
        if t_frac < 0.7: decay_factor = 1 - t_frac
        else: decay_factor = 0.3 * np.exp(-5 * (t_frac - 0.7))
        return initial_theta * decay_factor
    else:
        decay_factor = np.exp(-2 * t_frac)
        return initial_theta * (1 - decay_factor)

def get_trade_lifecycle_data(row, snapshots_df):
    days = int(row['Days Held'])
    if days < 1: days = 1
    total_pnl = row['P&L']
    
    if snapshots_df is not None and not snapshots_df.empty:
        trade_snaps = snapshots_df[snapshots_df['trade_id'] == row['id']].sort_values('days_held').copy()
        if len(trade_snaps) >= 2:
            trade_snaps['Cumulative_PnL'] = trade_snaps['pnl']
            trade_snaps['Day'] = trade_snaps['days_held']
            if trade_snaps['Day'].min() > 0:
                start_row = pd.DataFrame({'Day': [0], 'Cumulative_PnL': [0]})
                trade_snaps = pd.concat([start_row, trade_snaps], ignore_index=True)
            if row['Status'] == 'Expired':
                last_snap_day = trade_snaps['Day'].max()
                if last_snap_day < days:
                    end_row = pd.DataFrame({'Day': [days], 'Cumulative_PnL': [total_pnl]})
                    trade_snaps = pd.concat([trade_snaps, end_row], ignore_index=True)
            trade_snaps['Pct_Duration'] = (trade_snaps['Day'] / days) * 100
            denom = abs(total_pnl) if abs(total_pnl) > 0 else 1
            trade_snaps['Pct_PnL'] = (trade_snaps['Cumulative_PnL'] / denom) * 100
            return trade_snaps[['Day', 'Cumulative_PnL', 'Pct_Duration', 'Pct_PnL']]

    daily_data = []
    initial_theta = row['Theta'] if row['Theta'] != 0 else 1.0
    weights = []
    for d in range(1, days + 1):
        w = theta_decay_model(initial_theta, d, row['Strategy'], max(45, days))
        weights.append(abs(w))
    
    total_w = sum(weights)
    if total_w == 0: weights = [1/days] * days
    else: weights = [w/total_w for w in weights]
    
    cum_pnl = 0
    daily_data.append({'Day': 0, 'Cumulative_PnL': 0, 'Pct_Duration': 0, 'Pct_PnL': 0})
    for i, w in enumerate(weights):
        day_num = i + 1
        day_gain = total_pnl * w
        cum_pnl += day_gain
        pct_dur = (day_num / days) * 100
        denom = abs(total_pnl) if abs(total_pnl) > 0 else 1
        pct_pnl = (cum_pnl / denom) * 100
        daily_data.append({'Day': day_num, 'Cumulative_PnL': cum_pnl, 'Pct_Duration': pct_dur, 'Pct_PnL': pct_pnl})
    return pd.DataFrame(daily_data)

def reconstruct_daily_pnl(trades_df):
    trades = trades_df.copy()
    trades['Entry Date'] = pd.to_datetime(trades['Entry Date'])
    trades['Exit Date'] = pd.to_datetime(trades['Exit Date'])
    start_date = trades['Entry Date'].min()
    end_date = max(trades['Exit Date'].max(), pd.Timestamp.now())
    date_range = pd.date_range(start=start_date, end=end_date)
    daily_pnl_dict = {d.date(): 0.0 for d in date_range}

    for _, trade in trades.iterrows():
        if pd.isnull(trade['Exit Date']): continue
        days = trade['Days Held']
        if days <= 0: days = 1
        total_pnl = trade['P&L']
        strategy = trade['Strategy']
        initial_theta = trade['Theta'] if trade['Theta'] != 0 else 1.0
        
        daily_theta_weights = []
        for day in range(days):
            expected_theta = theta_decay_model(initial_theta, day, strategy, max(45, days))
            daily_theta_weights.append(abs(expected_theta))

        total_theta_sum = sum(daily_theta_weights)
        if total_theta_sum == 0: daily_theta_weights = [1/days] * days
        else: daily_theta_weights = [w/total_theta_sum for w in daily_theta_weights]
            
        curr = trade['Entry Date']
        for day_weight in daily_theta_weights:
            if curr.date() in daily_pnl_dict:
                daily_pnl_dict[curr.date()] += total_pnl * day_weight
            else:
                daily_pnl_dict[curr.date()] = total_pnl * day_weight
            curr += pd.Timedelta(days=1)
    return daily_pnl_dict

# --- SMART FILE PARSER ---
def parse_optionstrat_file(file, file_type, config_dict):
    try:
        df_raw = None
        if file.name.endswith(('.xlsx', '.xls')):
            try:
                df_temp = pd.read_excel(file, header=None)
                header_row = 0
                for i, row in df_temp.head(30).iterrows():
                    row_vals = [str(v).strip() for v in row.values]
                    if "Name" in row_vals and "Total Return $" in row_vals:
                        header_row = i
                        break
                file.seek(0)
                df_raw = pd.read_excel(file, header=header_row)
                
                if 'Link' in df_raw.columns:
                    try:
                        file.seek(0)
                        wb = load_workbook(file, data_only=False)
                        sheet = wb.active
                        excel_header_row = header_row + 1
                        link_col_idx = None
                        for cell in sheet[excel_header_row]:
                            if str(cell.value).strip() == "Link":
                                link_col_idx = cell.col_idx
                                break
                        if link_col_idx:
                            links = []
                            for i in range(len(df_raw)):
                                excel_row_idx = excel_header_row + 1 + i
                                cell = sheet.cell(row=excel_row_idx, column=link_col_idx)
                                url = ""
                                if cell.hyperlink: url = cell.hyperlink.target
                                elif cell.value and str(cell.value).startswith('=HYPERLINK'):
                                    try:
                                        parts = str(cell.value).split('"')
                                        if len(parts) > 1: url = parts[1]
                                    except: pass
                                links.append(url if url else "")
                            df_raw['Link'] = links
                    except: pass
            except: pass

        if df_raw is None:
            file.seek(0)
            content = file.getvalue().decode("utf-8", errors='ignore')
            lines = content.split('\n')
            header_row = 0
            for i, line in enumerate(lines[:30]):
                if "Name" in line and "Total Return" in line:
                    header_row = i
                    break
            file.seek(0)
            df_raw = pd.read_csv(file, skiprows=header_row)

        parsed_trades = []
        current_trade = None
        current_legs = []

        def finalize_trade(trade_data, legs, f_type):
            if not trade_data.any(): return None
            
            name = str(trade_data.get('Name', ''))
            group = str(trade_data.get('Group', '')) 
            created = trade_data.get('Created At', '')
            try: start_dt = pd.to_datetime(created)
            except: return None 

            strat = get_strategy_dynamic(name, group, config_dict)
            link = str(trade_data.get('Link', ''))
            if link == 'nan' or link == 'Open': link = "" 
            
            pnl = clean_num(trade_data.get('Total Return $', 0))
            debit = abs(clean_num(trade_data.get('Net Debit/Credit', 0)))
            theta = clean_num(trade_data.get('Theta', 0))
            delta = clean_num(trade_data.get('Delta', 0))
            gamma = clean_num(trade_data.get('Gamma', 0))
            vega = clean_num(trade_data.get('Vega', 0))
            iv = clean_num(trade_data.get('IV', 0))

            exit_dt = None
            try:
                raw_exp = trade_data.get('Expiration')
                if pd.notnull(raw_exp) and str(raw_exp).strip() != '':
                    exit_dt = pd.to_datetime(raw_exp)
            except: pass

            days_held = 1
            if exit_dt and f_type == "History": days_held = (exit_dt - start_dt).days
            else: days_held = (datetime.now() - start_dt).days
            if days_held < 1: days_held = 1

            strat_config = config_dict.get(strat, {})
            typical_debit = strat_config.get('debit_per_lot', 5000)
            
            lot_match = re.search(r'(\d+)\s*(?:LOT|L\b)', name, re.IGNORECASE)
            if lot_match: lot_size = int(lot_match.group(1))
            else: lot_size = int(round(debit / typical_debit))
            if lot_size < 1: lot_size = 1

            put_pnl = 0.0
            call_pnl = 0.0
            
            if f_type == "History":
                for leg in legs:
                    if len(leg) < 5: continue
                    sym = str(leg.iloc[0]) 
                    if not sym.startswith('.'): continue
                    try:
                        qty = clean_num(leg.iloc[1])
                        entry = clean_num(leg.iloc[2])
                        close_price = clean_num(leg.iloc[4])
                        leg_pnl = (close_price - entry) * qty * 100
                        if 'P' in sym and 'C' not in sym: put_pnl += leg_pnl
                        elif 'C' in sym and 'P' not in sym: call_pnl += leg_pnl
                        elif re.search(r'[0-9]P[0-9]', sym): put_pnl += leg_pnl
                        elif re.search(r'[0-9]C[0-9]', sym): call_pnl += leg_pnl
                    except: pass
            
            t_id = generate_id(name, strat, start_dt)
            return {
                'id': t_id, 'name': name, 'strategy': strat, 'start_dt': start_dt,
                'exit_dt': exit_dt, 'days_held': days_held, 'debit': debit,
                'lot_size': lot_size, 'pnl': pnl, 
                'theta': theta, 'delta': delta, 'gamma': gamma, 'vega': vega,
                'iv': iv, 'put_pnl': put_pnl, 'call_pnl': call_pnl, 'link': link,
                'group': group
            }

        cols = df_raw.columns
        col_names = [str(c) for c in cols]
        if 'Name' not in col_names or 'Total Return $' not in col_names:
            return []

        for index, row in df_raw.iterrows():
            name_val = str(row['Name'])
            if name_val and not name_val.startswith('.') and name_val != 'Symbol' and name_val != 'nan':
                if current_trade is not None:
                    res = finalize_trade(current_trade, current_legs, file_type)
                    if res: parsed_trades.append(res)
                current_trade = row
                current_legs = []
            elif name_val.startswith('.'):
                current_legs.append(row)
        
        if current_trade is not None:
             res = finalize_trade(current_trade, current_legs, file_type)
             if res: parsed_trades.append(res)
        return parsed_trades
    except Exception as e:
        print(f"Parser Error: {e}")
        return []

# --- SYNC ENGINE ---
def sync_data(file_list, file_type):
    log = []
    if not isinstance(file_list, list): file_list = [file_list]
    conn = get_db_connection()
    c = conn.cursor()
    db_active_ids = set()
    if file_type == "Active":
        try:
            current_active = pd.read_sql("SELECT id FROM trades WHERE status = 'Active'", conn)
            db_active_ids = set(current_active['id'].tolist())
        except: pass
    file_found_ids = set()
    
    config_dict = load_strategy_config()

    for file in file_list:
        count_new = 0
        count_update = 0
        try:
            trades_data = parse_optionstrat_file(file, file_type, config_dict)
            if not trades_data:
                log.append(f" {file.name}: Skipped (No valid trades found)")
                continue

            for t in trades_data:
                trade_id = t['id']
                if file_type == "Active":
                    file_found_ids.add(trade_id)
                
                c.execute("SELECT id, status, theta, delta, gamma, vega, put_pnl, call_pnl, iv, link, lot_size, strategy FROM trades WHERE id = ?", (trade_id,))
                existing = c.fetchone()
                
                if existing is None and t['link'] and len(t['link']) > 15:
                    c.execute("SELECT id, name FROM trades WHERE link = ?", (t['link'],))
                    link_match = c.fetchone()
                    if link_match:
                        old_id, old_name = link_match
                        try:
                            c.execute("UPDATE snapshots SET trade_id = ? WHERE trade_id = ?", (trade_id, old_id))
                            c.execute("UPDATE trades SET id=?, name=? WHERE id=?", (trade_id, t['name'], old_id))
                            log.append(f" Renamed: '{old_name}' -> '{t['name']}'")
                            c.execute("SELECT id, status, theta, delta, gamma, vega, put_pnl, call_pnl, iv, link, lot_size, strategy FROM trades WHERE id = ?", (trade_id,))
                            existing = c.fetchone()
                            if file_type == "Active":
                                file_found_ids.add(trade_id)
                                if old_id in db_active_ids: db_active_ids.remove(old_id)
                                db_active_ids.add(trade_id)
                        except Exception as rename_err:
                            print(f"Rename failed: {rename_err}")

                status = "Active" if file_type == "Active" else "Expired"
                
                if existing is None:
                    c.execute('''INSERT INTO trades 
                        (id, name, strategy, status, entry_date, exit_date, days_held, debit, lot_size, pnl, theta, delta, gamma, vega, notes, tags, parent_id, put_pnl, call_pnl, iv, link, original_group)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (trade_id, t['name'], t['strategy'], status, t['start_dt'].date(), 
                         t['exit_dt'].date() if t['exit_dt'] else None, 
                         t['days_held'], t['debit'], t['lot_size'], t['pnl'], 
                         t['theta'], t['delta'], t['gamma'], t['vega'], "", "", "", t['put_pnl'], t['call_pnl'], t['iv'], t['link'], t['group']))
                    count_new += 1
                else:
                    db_lot_size = existing[10]
                    final_lot_size = t['lot_size']
                    if db_lot_size and db_lot_size > 0:
                        final_lot_size = db_lot_size

                    db_strategy = existing[11]
                    final_strategy = db_strategy
                    if db_strategy == 'Other' and t['strategy'] != 'Other':
                         final_strategy = t['strategy']

                    old_put = existing[6] if existing[6] else 0.0
                    old_call = existing[7] if existing[7] else 0.0
                    old_iv = existing[8] if existing[8] else 0.0
                    old_link = existing[9] if existing[9] else ""
                    
                    old_status = existing[1]
                    old_theta = existing[2]

                    final_theta = t['theta'] if t['theta'] != 0 else old_theta
                    final_delta = t['delta'] if t['delta'] != 0 else 0
                    final_gamma = t['gamma'] if t['gamma'] != 0 else 0
                    final_vega = t['vega'] if t['vega'] != 0 else 0
                    final_iv = t['iv'] if t['iv'] != 0 else old_iv
                    final_put = t['put_pnl'] if t['put_pnl'] != 0 else old_put
                    final_call = t['call_pnl'] if t['call_pnl'] != 0 else old_call
                    final_link = t['link'] if t['link'] != "" else old_link

                    if file_type == "History":
                        c.execute('''UPDATE trades SET 
                            pnl=?, status=?, exit_date=?, days_held=?, theta=?, delta=?, gamma=?, vega=?, put_pnl=?, call_pnl=?, iv=?, link=?, lot_size=?, strategy=?, original_group=?
                            WHERE id=?''', 
                            (t['pnl'], status, t['exit_dt'].date() if t['exit_dt'] else None, t['days_held'], 
                             final_theta, final_delta, final_gamma, final_vega, final_put, final_call, final_iv, final_link, final_lot_size, final_strategy, t['group'], trade_id))
                        count_update += 1
                    elif old_status in ["Active", "Missing"]: 
                        c.execute('''UPDATE trades SET 
                            pnl=?, days_held=?, theta=?, delta=?, gamma=?, vega=?, iv=?, link=?, status='Active', exit_date=?, lot_size=?, strategy=?, original_group=?
                            WHERE id=?''', 
                            (t['pnl'], t['days_held'], final_theta, final_delta, final_gamma, final_vega, final_iv, final_link, 
                             t['exit_dt'].date() if t['exit_dt'] else None, final_lot_size, final_strategy, t['group'], trade_id))
                        count_update += 1
                if file_type == "Active":
                    today = datetime.now().date()
                    c.execute("SELECT id FROM snapshots WHERE trade_id=? AND snapshot_date=?", (trade_id, today))
                    
                    theta_val = t['theta'] if t['theta'] else 0.0
                    delta_val = t['delta'] if t['delta'] else 0.0
                    vega_val = t['vega'] if t['vega'] else 0.0
                    gamma_val = t['gamma'] if t['gamma'] else 0.0
                    
                    if not c.fetchone():
                        c.execute("INSERT INTO snapshots (trade_id, snapshot_date, pnl, days_held, theta, delta, vega, gamma) VALUES (?,?,?,?,?,?,?,?)",
                                  (trade_id, today, t['pnl'], t['days_held'], theta_val, delta_val, vega_val, gamma_val))
                    else:
                        c.execute("UPDATE snapshots SET pnl=?, days_held=?, theta=?, delta=?, vega=?, gamma=? WHERE trade_id=? AND snapshot_date=?",
                                  (t['pnl'], t['days_held'], theta_val, delta_val, vega_val, gamma_val, trade_id, today))
            log.append(f" {file.name}: {count_new} New, {count_update} Updated")
        except Exception as e:
            log.append(f" {file.name}: Error - {str(e)}")
            
    if file_type == "Active":
        missing_ids = db_active_ids - file_found_ids
        if missing_ids:
            placeholders = ','.join('?' for _ in missing_ids)
            c.execute(f"UPDATE trades SET status = 'Missing' WHERE id IN ({placeholders})", list(missing_ids))
            log.append(f" Integrity: Marked {len(missing_ids)} trades as 'Missing'.")
    conn.commit()
    conn.close()
    return log

def update_journal(edited_df):
    conn = get_db_connection()
    c = conn.cursor()
    count = 0
    try:
        for index, row in edited_df.iterrows():
            t_id = row['id'] 
            notes = str(row['Notes'])
            tags = str(row['Tags'])
            pid = str(row['Parent ID'])
            new_lot = int(row['lot_size']) if 'lot_size' in row and row['lot_size'] > 0 else 1
            new_strat = str(row['Strategy']) 
            
            c.execute("UPDATE trades SET notes=?, tags=?, parent_id=?, lot_size=?, strategy=? WHERE id=?", (notes, tags, pid, new_lot, new_strat, t_id))
            count += 1
        conn.commit()
        return count
    except Exception as e: return 0
    finally: conn.close()

def update_strategy_config(edited_df):
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM strategy_config")
        for i, row in edited_df.iterrows():
            c.execute("INSERT INTO strategy_config VALUES (?,?,?,?,?,?,?)", 
                      (row['Name'], row['Identifier'], row['Target PnL'], row['Target Days'], row['Min Stability'], row['Description'], row['Typical Debit']))
        conn.commit()
        return True
    except Exception as e:
        print(e)
        return False
    finally: conn.close()

def reprocess_other_trades():
    conn = get_db_connection()
    c = conn.cursor()
    config_dict = load_strategy_config()
    try:
        c.execute("SELECT id, name, original_group, strategy FROM trades")
    except:
        c.execute("SELECT id, name, '', strategy FROM trades")
    all_trades = c.fetchall()
    updated_count = 0
    for t_id, t_name, t_group, current_strat in all_trades:
        if current_strat == "Other":
            group_val = t_group if t_group else ""
            new_strat = get_strategy_dynamic(t_name, group_val, config_dict) 
            if new_strat != "Other":
                c.execute("UPDATE trades SET strategy = ? WHERE id = ?", (new_strat, t_id))
                updated_count += 1
    conn.commit()
    conn.close()
    return updated_count

def get_trade_group_series(trades_df):
    if trades_df is None or trades_df.empty:
        return pd.Series(dtype=str)

    index = trades_df.index
    if 'original_group' in trades_df.columns:
        group_series = trades_df['original_group'].fillna('').astype(str).str.strip()
    else:
        group_series = pd.Series([''] * len(trades_df), index=index, dtype=str)

    group_series = group_series.replace({'nan': '', 'None': ''})

    if 'Strategy' in trades_df.columns:
        fallback = trades_df['Strategy'].fillna('').astype(str).str.strip()
    else:
        fallback = pd.Series([''] * len(trades_df), index=index, dtype=str)

    group_series = group_series.where(group_series != '', fallback)
    group_series = group_series.replace({'nan': 'Uncategorized', 'None': 'Uncategorized', '': 'Uncategorized'})
    return group_series

def is_test_group(group_name):
    text = str(group_name or '').strip().lower()
    return any(token in text for token in ['test', 'demo', 'paper', 'sandbox', 'tmp'])

def get_group_universe_options(trades_df):
    if trades_df is None or trades_df.empty:
        return []
    groups = get_trade_group_series(trades_df).dropna().astype(str).str.strip()
    groups = [g for g in groups.tolist() if g and g != 'nan']
    return sorted(dict.fromkeys(groups))

def apply_group_universe_filter(trades_df):
    if trades_df is None or trades_df.empty:
        return trades_df
    excluded = set(st.session_state.get('group_universe_excluded', []))
    if not excluded:
        return trades_df.copy()
    group_series = get_trade_group_series(trades_df)
    return trades_df.loc[~group_series.isin(excluded)].copy()

def render_group_universe_control(trades_df):
    groups = get_group_universe_options(trades_df)
    if 'group_universe_excluded' not in st.session_state:
        st.session_state['group_universe_excluded'] = []

    with st.sidebar.expander("Group Universe Control", expanded=True):
        if not groups:
            st.caption("Upload both workbooks to manage group filters.")
            return []

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Include Everything", key="group_universe_include_all"):
                st.session_state['group_universe_excluded'] = []
        with c2:
            if st.button("Exclude Test Groups", key="group_universe_exclude_test"):
                st.session_state['group_universe_excluded'] = [g for g in groups if is_test_group(g)]

        excluded = st.multiselect(
            "Exclude groups from calculations",
            options=groups,
            default=st.session_state.get('group_universe_excluded', []),
            key="group_universe_excluded",
        )
        st.caption(
            f"Included groups: {len(groups) - len(set(excluded))} • "
            f"Excluded groups: {len(set(excluded))} • "
            f"Active calculation universe: {int((~get_trade_group_series(trades_df).isin(set(excluded))).sum())} trades"
        )
        return excluded

# --- DATA LOADER ---
@st.cache_data(ttl=60)
def load_data():
    empty_schema = pd.DataFrame(columns=[
        'id', 'Name', 'Strategy', 'Status', 'P&L', 'Debit', 
        'Days Held', 'Theta', 'Delta', 'Gamma', 'Vega', 
        'Entry Date', 'Exit Date', 'Notes', 'Tags', 
        'Parent ID', 'Put P&L', 'Call P&L', 'IV', 'Link',
        'lot_size', 'Debit/Lot', 'ROI', 'Daily Yield %', 
        'Ann. ROI', 'Theta Pot.', 'Theta Eff.', 
        'Theta/Cap %', 'Ticker', 'Stability', 'Grade', 'Reason', 'P&L Vol'
    ])
    
    if not os.path.exists(DB_NAME): return empty_schema
    conn = get_db_connection()
    try:
        df = pd.read_sql("SELECT * FROM trades", conn)
        if df.empty: return empty_schema
        
        snaps = pd.read_sql("SELECT trade_id, pnl FROM snapshots", conn)
        if not snaps.empty:
            vol_df = snaps.groupby('trade_id')['pnl'].std().reset_index()
            vol_df.rename(columns={'pnl': 'P&L Vol'}, inplace=True)
            df = df.merge(vol_df, left_on='id', right_on='trade_id', how='left')
            df['P&L Vol'] = df['P&L Vol'].fillna(0)
        else: df['P&L Vol'] = 0.0
    except Exception as e: return empty_schema
    finally: conn.close()
    
    if not df.empty:
        df = df.rename(columns={
            'name': 'Name', 'strategy': 'Strategy', 'status': 'Status',
            'pnl': 'P&L', 'debit': 'Debit', 'days_held': 'Days Held',
            'theta': 'Theta', 'delta': 'Delta', 'gamma': 'Gamma', 'vega': 'Vega',
            'entry_date': 'Entry Date', 'exit_date': 'Exit Date', 'notes': 'Notes',
            'tags': 'Tags', 'parent_id': 'Parent ID', 
            'put_pnl': 'Put P&L', 'call_pnl': 'Call P&L', 'iv': 'IV', 'link': 'Link'
        })
        
        required_cols = ['Gamma', 'Vega', 'Theta', 'Delta', 'P&L', 'Debit', 'lot_size', 'Notes', 'Tags', 'Parent ID', 'Put P&L', 'Call P&L', 'IV', 'Link']
        for col in required_cols:
            if col not in df.columns: df[col] = "" if col in ['Notes', 'Tags', 'Parent ID', 'Link'] else 0.0
        
        numeric_cols = ['Debit', 'P&L', 'Days Held', 'Theta', 'Delta', 'Gamma', 'Vega', 'IV', 'Put P&L', 'Call P&L']
        for c in numeric_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        df['Entry Date'] = pd.to_datetime(df['Entry Date'])
        df['Exit Date'] = pd.to_datetime(df['Exit Date'])
        
        df['lot_size'] = pd.to_numeric(df['lot_size'], errors='coerce').fillna(1).astype(int)
        df['lot_size'] = df['lot_size'].apply(lambda x: 1 if x < 1 else x)
        
        df['Debit/Lot'] = np.where(df['lot_size'] > 0, df['Debit'] / df['lot_size'], df['Debit'])

        df['ROI'] = (df['P&L'] / df['Debit'].replace(0, 1) * 100)
        df['Daily Yield %'] = np.where(df['Days Held'] > 0, df['ROI'] / df['Days Held'], 0)
        df['Ann. ROI'] = df['Daily Yield %'] * 365
        df['Theta Pot.'] = df['Theta'] * df['Days Held']
        
        # V150: Vectorized Theta Efficiency logic applied
        df = vectorized_theta_efficiency(df)
        df['Theta Eff.'] = df.get('Theta_Efficiency_Score', np.where(df['Theta Pot.'] > 0, df['P&L'] / df['Theta Pot.'], 0.0))
        
        df['Theta/Cap %'] = np.where(df['Debit'] > 0, (df['Theta'] / df['Debit']) * 100, 0)
        df['Ticker'] = df['Name'].apply(extract_ticker)
        
        df['Parent ID'] = df['Parent ID'].astype(str).str.strip().replace('nan', '').replace('None', '')
        
        # V150: Improved stability formula
        df['Stability'] = df.get('Stability_Score', np.where(df['Theta'] > 0, df['Theta'] / (df['Delta'].abs() + 1), 0.0))
        
        def get_grade(row):
            s, d = row['Strategy'], row['Debit/Lot']
            reason = "Standard"
            grade = "C"
            if s == '130/160':
                if d > 4800: grade="F"; reason="Overpriced (> $4.8k)"
                elif 3500 <= d <= 4500: grade="A+"; reason="Sweet Spot"
                else: grade="B"; reason="Acceptable"
            elif s == '160/190':
                if 4800 <= d <= 5500: grade="A"; reason="Ideal Pricing"
                else: grade="C"; reason="Check Pricing"
            elif s == 'M200':
                if 7500 <= d <= 8500: grade, reason = "A", "Perfect Entry"
                else: grade, reason = "B", "Variance"
            elif s == 'SMSF':
                if d > 15000: grade="B"; reason="High Debit" 
                else: grade="A"; reason="Standard"
            return pd.Series([grade, reason])

        df[['Grade', 'Reason']] = df.apply(get_grade, axis=1)
    return df

@st.cache_data(ttl=300)
def load_snapshots():
    if not os.path.exists(DB_NAME): return pd.DataFrame()
    conn = get_db_connection()
    try:
        q = """
        SELECT s.snapshot_date, s.pnl, s.days_held, s.theta, s.delta, s.vega, s.gamma,
               t.strategy, t.name, t.id as trade_id, t.theta as initial_theta
        FROM snapshots s
        JOIN trades t ON s.trade_id = t.id
        """
        try:
            df = pd.read_sql(q, conn)
        except:
            q_fallback = """
            SELECT s.snapshot_date, s.pnl, s.days_held, s.theta, s.delta, s.vega, 
                   t.strategy, t.name, t.id as trade_id, t.theta as initial_theta
            FROM snapshots s
            JOIN trades t ON s.trade_id = t.id
            """
            df = pd.read_sql(q_fallback, conn)
            df['gamma'] = 0.0

        df['snapshot_date'] = pd.to_datetime(df['snapshot_date'])
        df['pnl'] = pd.to_numeric(df['pnl'], errors='coerce').fillna(0)
        df['days_held'] = pd.to_numeric(df['days_held'], errors='coerce').fillna(0)
        df['theta'] = pd.to_numeric(df['theta'], errors='coerce').fillna(0)
        df['delta'] = pd.to_numeric(df['delta'], errors='coerce').fillna(0)
        df['vega'] = pd.to_numeric(df['vega'], errors='coerce').fillna(0)
        df['gamma'] = pd.to_numeric(df['gamma'], errors='coerce').fillna(0)
        df['initial_theta'] = pd.to_numeric(df['initial_theta'], errors='coerce').fillna(0)
        return df
    except: return pd.DataFrame()
    finally: conn.close()

# --- INTELLIGENCE FUNCTIONS ---
def get_dynamic_targets(history_df, percentile):
    if history_df.empty: return {}
    winners = history_df[history_df['P&L'] > 0]
    if winners.empty: return {}
    targets = {}
    for strat, grp in winners.groupby('Strategy'):
        targets[strat] = {
            'Median Win': grp['P&L'].median(),
            'Optimal Exit': grp['P&L'].quantile(percentile)
        }
    return targets

def find_similar_trades(current_trade, historical_df, top_n=3):
    if historical_df.empty: return pd.DataFrame()
    features = ['Theta/Cap %', 'Delta', 'Debit/Lot']
    for f in features:
        if f not in current_trade or f not in historical_df.columns:
            return pd.DataFrame()
    curr_vec = np.nan_to_num(current_trade[features].values.astype(float)).reshape(1, -1)
    hist_vecs = np.nan_to_num(historical_df[features].values.astype(float))
    distances = cdist(curr_vec, hist_vecs, metric='euclidean')[0]
    similar_idx = np.argsort(distances)[:top_n]
    similar = historical_df.iloc[similar_idx].copy()
    max_dist = distances.max() if distances.max() > 0 else 1
    similar['Similarity %'] = 100 * (1 - distances[similar_idx] / max_dist)
    return similar[['Name', 'P&L', 'Days Held', 'ROI', 'Similarity %']]

def calculate_portfolio_metrics(trades_df, capital):
    if trades_df.empty or capital <= 0: return 0.0, 0.0
    daily_pnl_dict = reconstruct_daily_pnl(trades_df)
    trades_df['Entry Date'] = pd.to_datetime(trades_df['Entry Date'])
    trades_df['Exit Date'] = pd.to_datetime(trades_df['Exit Date'])
    start_date = trades_df['Entry Date'].min()
    end_date = max(trades_df['Exit Date'].max(), pd.Timestamp.now())
    date_range = pd.date_range(start=start_date, end=end_date)
    equity = capital
    daily_equity_values = []
    for d in date_range:
        day_pnl = daily_pnl_dict.get(d.date(), 0)
        equity += day_pnl
        daily_equity_values.append(equity)
    equity_series = pd.Series(daily_equity_values)
    daily_returns = equity_series.pct_change().dropna()
    if daily_returns.std() == 0:
        sharpe = 0.0
    else:
        sharpe = (daily_returns.mean() / daily_returns.std()) * np.sqrt(252)
    total_days = (end_date - start_date).days
    if total_days < 1: total_days = 1
    total_pnl = trades_df['P&L'].sum()
    end_val = capital + total_pnl
    try:
        cagr = ( (end_val / capital) ** (365 / total_days) ) - 1
    except:
        cagr = 0.0
    return sharpe, cagr * 100

def check_concentration_risk(active_df, total_equity, threshold=0.15):
    if active_df.empty or total_equity <= 0: return pd.DataFrame()
    warnings = []
    for _, row in active_df.iterrows():
        concentration = row['Debit'] / total_equity
        if concentration > threshold:
            warnings.append({
                'Trade': row['Name'], 'Strategy': row['Strategy'], 'Size %': f"{concentration:.1%}",
                'Risk': f"${row['Debit']:,.0f}", 'Limit': f"{threshold:.0%}"
            })
    return pd.DataFrame(warnings)

def calculate_max_drawdown(trades_df, initial_capital):
    if trades_df.empty or initial_capital <= 0: 
        return {'Max Drawdown %': 0.0, 'Current DD %': 0.0}
    daily_pnl_dict = reconstruct_daily_pnl(trades_df)
    trades_df['Entry Date'] = pd.to_datetime(trades_df['Entry Date'])
    trades_df['Exit Date'] = pd.to_datetime(trades_df['Exit Date'])
    start_date = trades_df['Entry Date'].min()
    end_date = max(trades_df['Exit Date'].max(), pd.Timestamp.now())
    date_range = pd.date_range(start=start_date, end=end_date)
    equity = initial_capital
    equity_curve = []
    dates = []
    for d in date_range:
        day_pnl = daily_pnl_dict.get(d.date(), 0)
        equity += day_pnl
        equity_curve.append(equity)
        dates.append(d.date())
    equity_series = pd.Series(equity_curve, index=pd.to_datetime(dates))
    running_max = equity_series.cummax()
    drawdown = (equity_series - running_max) / running_max
    max_dd = drawdown.min()
    current_dd = drawdown.iloc[-1]
    return {'Max Drawdown %': max_dd * 100, 'Current DD %': current_dd * 100}

def build_open_risk_snapshot(active_df):
    if active_df is None or active_df.empty:
        return {'pnl': 0.0, 'delta': 0.0, 'theta': 0.0, 'gamma': 0.0, 'vega': 0.0, 'count': 0}
    return {
        'pnl': float(active_df['P&L'].sum()),
        'delta': float(active_df['Delta'].sum()),
        'theta': float(active_df['Theta'].sum()),
        'gamma': float(active_df['Gamma'].sum()),
        'vega': float(active_df['Vega'].sum()),
        'count': int(len(active_df))
    }

def build_risk_limits(open_risk):
    abs_delta = abs(float(open_risk.get('delta', 0.0)) or 0.0)
    abs_theta = abs(float(open_risk.get('theta', 0.0)) or 0.0)
    abs_gamma = abs(float(open_risk.get('gamma', 0.0)) or 0.0)
    abs_vega = abs(float(open_risk.get('vega', 0.0)) or 0.0)
    abs_open_pnl = abs(float(open_risk.get('pnl', 0.0)) or 0.0)
    position_count = int(open_risk.get('count', 0) or 0)

    delta_limit = max(100, int(np.ceil(abs_delta * 1.35 / 25.0) * 25))
    theta_limit = max(25, int(np.ceil(abs_theta * 1.35 / 5.0) * 5))
    gamma_limit = max(1, float(np.ceil(abs_gamma * 1.5 * 100) / 100))
    vega_limit = max(50, int(np.ceil(abs_vega * 1.35 / 10.0) * 10))
    pnl_limit = max(1000, int(np.ceil(abs_open_pnl * 1.5 / 250.0) * 250))
    count_limit = max(10, int(np.ceil(position_count * 1.3 / 5.0) * 5))

    items = [
        {'label': 'Delta Limit', 'current': abs_delta, 'signed_current': float(open_risk.get('delta', 0.0) or 0.0), 'limit': delta_limit, 'unit': 'Δ', 'formatter': lambda v: f"{v:.2f}"},
        {'label': 'Theta Limit', 'current': abs_theta, 'signed_current': float(open_risk.get('theta', 0.0) or 0.0), 'limit': theta_limit, 'unit': 'Θ', 'formatter': lambda v: f"{v:.2f}"},
        {'label': 'Gamma Limit', 'current': abs_gamma, 'signed_current': float(open_risk.get('gamma', 0.0) or 0.0), 'limit': gamma_limit, 'unit': 'Γ', 'formatter': lambda v: f"{v:.3f}"},
        {'label': 'Vega Limit', 'current': abs_vega, 'signed_current': float(open_risk.get('vega', 0.0) or 0.0), 'limit': vega_limit, 'unit': 'V', 'formatter': lambda v: f"{v:.2f}"},
        {'label': 'Open PnL Limit', 'current': abs_open_pnl, 'signed_current': float(open_risk.get('pnl', 0.0) or 0.0), 'limit': pnl_limit, 'unit': '$', 'formatter': lambda v: f"${v:,.0f}"},
        {'label': 'Open Position Limit', 'current': float(position_count), 'signed_current': float(position_count), 'limit': count_limit, 'unit': '#', 'formatter': lambda v: f"{int(round(v))}"}
    ]

    for item in items:
        item['ratio'] = item['current'] / item['limit'] if item['limit'] else 0.0
        if item['ratio'] >= 0.9:
            item['status'] = 'Near limit'
        elif item['ratio'] >= 0.7:
            item['status'] = 'Watch'
        else:
            item['status'] = 'Comfortable'
    return items

def build_equity_drawdown_curve(expired_df, initial_capital):
    if expired_df is None or expired_df.empty or initial_capital <= 0:
        return pd.DataFrame()

    curve_df = expired_df.dropna(subset=['Exit Date']).copy()
    if curve_df.empty:
        return pd.DataFrame()

    curve_df['Exit Date'] = pd.to_datetime(curve_df['Exit Date']).dt.date
    daily_pnl = curve_df.groupby('Exit Date')['P&L'].sum().sort_index().reset_index()
    daily_pnl['Equity'] = initial_capital + daily_pnl['P&L'].cumsum()
    daily_pnl['Running Max'] = daily_pnl['Equity'].cummax()
    daily_pnl['Drawdown %'] = np.where(
        daily_pnl['Running Max'] > 0,
        ((daily_pnl['Equity'] - daily_pnl['Running Max']) / daily_pnl['Running Max']) * 100,
        0.0
    )
    return daily_pnl[['Exit Date', 'Equity', 'Drawdown %']]

def rolling_correlation_matrix(snaps, window_days=30):
    if snaps.empty: return None
    strat_daily = snaps.pivot_table(index='snapshot_date', columns='strategy', values='pnl', aggfunc='sum')
    if len(strat_daily) < window_days: return None
    last_30 = strat_daily.tail(30)
    corr_30 = last_30.corr()
    fig = px.imshow(corr_30, text_auto=".2f", aspect="auto", color_continuous_scale="RdBu", 
                    title="Strategy Correlation (Last 30 Days)", labels=dict(color="Correlation"))
    fig = apply_chart_theme(fig)
    return fig

def generate_adaptive_rulebook_text(history_df, strategies):
    text = "#  The Adaptive Trader's Constitution\n*Rules evolve. This book rewrites itself based on your actual data.*\n\n"
    if history_df.empty:
        text += " *Not enough data yet. Complete more trades to unlock adaptive rules.*"
        return text
    for strat in strategies:
        strat_df = history_df[history_df['Strategy'] == strat]
        if strat_df.empty: continue
        winners = strat_df[strat_df['P&L'] > 0]
        text += f"### {strat}\n"
        if not winners.empty:
            winners = winners.copy()
            winners['Day'] = winners['Entry Date'].dt.day_name()
            best_day = winners.groupby('Day')['P&L'].mean().idxmax()
            text += f"* ** Best Entry Day:** {best_day} (Highest Avg Win)\n"
            avg_hold = winners['Days Held'].mean()
            text += f"* ** Optimal Hold:** {avg_hold:.0f} Days (Avg Winner Duration)\n"
            avg_cost = winners['Debit/Lot'].mean()
            text += f"* ** Target Cost:** ${avg_cost:,.0f} (Avg Winner Debit per Lot)\n"
        losers = strat_df[strat_df['P&L'] < 0]
        if not losers.empty:
             avg_loss_hold = losers['Days Held'].mean()
             text += f"* ** Loss Pattern:** Losers held for avg {avg_loss_hold:.0f} days.\n"
        text += "\n"
    text += "---\n###  Universal AI Gates\n"
    text += "1. **Efficiency Check:** If 'Rot Detector' flags a trade, cut it. Your capital is stuck.\n"
    text += "2. **Probability Gate:** Check 'Win Prob %' before entering. If < 40%, skip even if the chart looks good.\n"
    return text

# --- INITIALIZE DB ---
init_db()

ui_df = load_data()

# --- SIDEBAR ---
st.sidebar.markdown("###  Daily Workflow")

if GOOGLE_DEPS_INSTALLED:
    with st.sidebar.expander("☁️ Cloud Sync (Google Drive)", expanded=True):
        if not drive_mgr.is_connected:
            st.error("No secrets found. Add 'gcp_service_account' to .streamlit/secrets.toml")
        else:
            c1, c2 = st.columns(2)
            with c1:
                if st.button("⬆️ Sync to Cloud"):
                    success, msg = drive_mgr.upload_db()
                    if not success and "CONFLICT" in msg:
                        st.error(msg)
                        st.session_state['show_conflict'] = True
                    elif success: 
                        st.success(msg)
                        st.session_state['show_conflict'] = False
                    else: st.error(msg)
            
            with c2:
                if st.button("⬇️ Pull from Cloud"):
                    success, msg = drive_mgr.download_db()
                    if not success and "CONFLICT" in msg:
                        st.error(msg)
                        st.session_state['show_pull_conflict'] = True
                    elif success: 
                        st.cache_data.clear()
                        st.success(msg)
                        st.rerun()
                    else: st.error(msg)
            
            # Conflict Resolution UI (Upload)
            if st.session_state.get('show_conflict'):
                st.warning("Upload Conflict:")
                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("⬇️ Pull (Safe)", key="res_pull"):
                        success, msg = drive_mgr.download_db()
                        if success:
                            st.success("Resolved: Cloud version pulled.")
                            st.session_state['show_conflict'] = False
                            st.cache_data.clear()
                            st.rerun()
                with cc2:
                    if st.button("⚠️ Force Push", key="res_force"):
                        success, msg = drive_mgr.upload_db(force=True)
                        if success:
                            st.success("Resolved: Cloud overwritten.")
                            st.session_state['show_conflict'] = False

            # Conflict Resolution UI (Pull)
            if st.session_state.get('show_pull_conflict'):
                st.warning("Pull Warning: Local file is newer!")
                cp1, cp2 = st.columns(2)
                with cp1:
                    if st.button("⬇️ Force Pull (Lose Local)", key="force_pull"):
                        success, msg = drive_mgr.download_db(force=True)
                        if success:
                            st.session_state['show_pull_conflict'] = False
                            st.cache_data.clear()
                            st.rerun()
                with cp2:
                    if st.button("❌ Cancel", key="cancel_pull"):
                        st.session_state['show_pull_conflict'] = False
                        st.rerun()
            
            # Persistent Status Indicator
            st.sidebar.divider()
            last_sync_time = st.session_state.get('last_cloud_sync')
            cloud_file_id, cloud_file_name = drive_mgr.find_db_file()
            if cloud_file_id:
                cloud_ts = drive_mgr.get_cloud_modified_time(cloud_file_id)
                if cloud_ts:
                    st.sidebar.caption(f"☁️ Cloud File: {cloud_ts.strftime('%b %d, %H:%M')} (UTC)")
            
            if last_sync_time:
                mins_ago = (datetime.now() - last_sync_time).total_seconds() / 60
                if mins_ago < 1:
                    st.sidebar.success(f"✅ Synced just now")
                elif mins_ago < 60:
                    st.sidebar.info(f"Last Sync: {int(mins_ago)}m ago")
                else:
                    st.sidebar.warning(f"Last Sync: {last_sync_time.strftime('%H:%M')}")
            else:
                st.sidebar.warning("⚠️ Session Status: Unsaved")

            with st.expander("🕵️ Debug: What can I see?"):
                st.write("Files visible to Bot:")
                files = drive_mgr.list_files_debug()
                if files:
                    for f in files:
                        mod_time = f['modifiedTime'] if 'modifiedTime' in f else 'Unknown'
                        st.code(f"{f['name']}\nID: {f['id']}\nMod: {mod_time}")
                else:
                    st.warning("No files found. Ensure you shared the DB with the service account email.")

with st.sidebar.expander("1.  STARTUP (Restore Local)", expanded=False):
    restore = st.file_uploader("Upload .db file", type=['db'], key='restore')
    if restore:
        with open(DB_NAME, "wb") as f: f.write(restore.getbuffer())
        st.cache_data.clear()
        st.success("Restored.")
        if 'restored' not in st.session_state:
            st.session_state['restored'] = True
            st.rerun()

st.sidebar.markdown(" *then...*")
with st.sidebar.expander("2.  WORK (Sync Files)", expanded=True):
    st.caption("Upload one workbook for all active trades and one workbook for all closed trades.")
    active_up = st.file_uploader("All Active Workbook", type=['xlsx', 'xls', 'csv'], key="act")
    history_up = st.file_uploader("All Closed Workbook", type=['xlsx', 'xls', 'csv'], key="hist")
    if st.button(" Process Workbooks"):
        logs = []
        if active_up: logs.extend(sync_data([active_up], "Active"))
        if history_up: logs.extend(sync_data([history_up], "History"))
        if logs:
            for l in logs: st.write(l)
            st.cache_data.clear()
            st.success("Sync Complete!")
            auto_sync_if_connected()

st.sidebar.markdown(" *finally...*")
with st.sidebar.expander("3.  SHUTDOWN (Local Backup)", expanded=False):
    with open(DB_NAME, "rb") as f:
        st.download_button(" Save Database File", f, "trade_guardian_v4.db", "application/x-sqlite3")

with st.sidebar.expander(" Maintenance", expanded=False):
    st.caption("Fix Duplicates / Rename Issues")
    if st.button(" Vacuum DB"):
        conn = get_db_connection()
        conn.execute("VACUUM")
        conn.close()
        st.success("Optimized.")
    st.markdown("---")
    conn = get_db_connection()
    try:
        all_trades = pd.read_sql("SELECT id, name, status, pnl, days_held FROM trades ORDER BY status, entry_date DESC", conn)
        if not all_trades.empty:
            st.write(" **Delete Specific Trades**")
            all_trades['Label'] = all_trades['name'] + " (" + all_trades['status'] + ", $" + all_trades['pnl'].astype(str) + ")"
            trades_to_del = st.multiselect("Select trades to delete:", all_trades['Label'].tolist())
            if st.button(" Delete Selected Trades"):
                if trades_to_del:
                    ids_to_del = all_trades[all_trades['Label'].isin(trades_to_del)]['id'].tolist()
                    placeholders = ','.join('?' for _ in ids_to_del)
                    conn.execute(f"DELETE FROM snapshots WHERE trade_id IN ({placeholders})", ids_to_del)
                    conn.execute(f"DELETE FROM trades WHERE id IN ({placeholders})", ids_to_del)
                    conn.commit()
                    st.success(f"Deleted {len(ids_to_del)} trades!")
                    st.cache_data.clear()
                    st.rerun()
    except: pass
    conn.close()
    st.markdown("---")
    if st.button(" Hard Reset (Delete All Data)"):
        conn = get_db_connection()
        conn.execute("DROP TABLE IF EXISTS trades")
        conn.execute("DROP TABLE IF EXISTS snapshots")
        conn.execute("DROP TABLE IF EXISTS strategy_config")
        conn.commit()
        conn.close()
        init_db()
        st.cache_data.clear()
        st.success("Wiped & Reset.")
        st.rerun()

st.sidebar.divider()
st.sidebar.header(" Portfolio Settings")

prime_cap = st.sidebar.number_input("Prime Account (130/160, M200)", min_value=1000, value=115000, step=1000)
smsf_cap = st.sidebar.number_input("SMSF Account", min_value=1000, value=150000, step=1000)
total_cap = prime_cap + smsf_cap

market_regime = st.sidebar.selectbox("Current Market Regime", ["Neutral (Standard)", "Bullish (Aggr. Targets)", "Bearish (Safe Targets)"], index=0)
regime_mult = 1.10 if "Bullish" in market_regime else 0.90 if "Bearish" in market_regime else 1.0

# --- SMART ADAPTIVE EXIT ENGINE ---
def calculate_decision_ladder(row, benchmarks_dict):
    strat = row['Strategy']
    days = row['Days Held']
    pnl = row['P&L']
    status = row['Status']
    theta = row['Theta']
    stability = row['Stability']
    debit = row['Debit']
    lot_size = row.get('lot_size', 1)
    if lot_size < 1: lot_size = 1
    juice_val = 0.0
    juice_type = "Neutral"
    if status == 'Missing': return "REVIEW", 100, "Missing from data", 0, "Error"
    bench = benchmarks_dict.get(strat, {})
    hist_avg_pnl = bench.get('pnl', 1000)
    target_profit = (hist_avg_pnl * regime_mult) * lot_size
    hist_avg_days = bench.get('dit', 40)
    score = 50 
    action = "HOLD"
    reason = "Normal"
    if pnl < 0:
        juice_type = "Recovery Days"
        if theta > 0:
            recov_days = abs(pnl) / theta
            juice_val = recov_days
            is_cooking = (strat == '160/190' and days < 30)
            is_young = days < 15
            if not is_cooking and not is_young:
                remaining_time_est = max(1, hist_avg_days - days)
                if recov_days > remaining_time_est:
                    score += 40
                    action = "STRUCTURAL FAILURE"
                    reason = f"Zombie (Recov {recov_days:.0f}d > Left {remaining_time_est:.0f}d)"
        else:
            juice_val = 999
            if days > 15:
                score += 30
                reason = "Negative Theta"
    else:
        juice_type = "Left in Tank"
        left_in_tank = max(0, target_profit - pnl)
        juice_val = left_in_tank
        if debit > 0 and (left_in_tank / debit) < 0.05:
            score += 40
            reason = "Squeezed Dry (Risk > Reward)"
        elif left_in_tank < (100 * lot_size):
            score += 35
            reason = f"Empty Tank (<${100*lot_size})"

    if pnl >= target_profit:
        return "TAKE PROFIT", 100, f"Hit Target ${target_profit:.0f}", juice_val, juice_type
    elif pnl >= target_profit * 0.8:
        score += 30
        action = "PREPARE EXIT"
        reason = "Near Target"
        
    stale_threshold = hist_avg_days * 1.25 
    if strat == '130/160':
        limit_130 = min(stale_threshold, 30) 
        if days > limit_130 and pnl < (100 * lot_size):
            return "KILL", 95, f"Stale (> {limit_130:.0f}d)", juice_val, juice_type
        elif days > (limit_130 * 0.8):
            score += 20
            reason = "Aging"
    elif strat == '160/190':
        cooking_limit = max(30, hist_avg_days * 0.7)
        if days < cooking_limit:
            score = 10 
            action = "COOKING" 
            reason = f"Too Early (<{cooking_limit:.0f}d)"
        elif days > stale_threshold:
            score += 25
            action = "WATCH"
            reason = f"Mature (>{stale_threshold:.0f}d)"
    elif strat == 'M200':
        if 13 <= days <= 15:
            score += 10
            action = "DAY 14 CHECK"
            reason = "Scheduled Review"
    if stability < 0.3 and days > 5:
        score += 25
        reason += " + Coin Flip (Unstable)"
        action = "RISK REVIEW"
    if row['Theta Eff.'] < 0.2 and days > 10:
        score += 15
        reason += " + Bad Decay"
    
    if pnl > 0 and days > 5:
        hist_vel = hist_avg_pnl / max(1, hist_avg_days)
        curr_vel = pnl / days
        if curr_vel < (hist_vel * 0.6):
             if pnl > (target_profit * 0.2):
                 score += 15
                 if action == "HOLD": 
                     action = "REDEPLOY?"
                     reason = f"Slow Capital (Vel ${curr_vel:.0f}/d vs Avg ${hist_vel:.0f}/d)"

    score = min(100, max(0, score))
    if score >= 90: action = "CRITICAL"
    elif score >= 70: action = "WATCH"
    elif score <= 30: action = "COOKING"
    return action, score, reason, juice_val, juice_type

# --- MAIN APP ---
render_group_universe_control(ui_df)
df = apply_group_universe_filter(load_data())

BASE_CONFIG = {
    '130/160': {'pnl': 500, 'dit': 36, 'stability': 0.8},
    '160/190': {'pnl': 700, 'dit': 44, 'stability': 0.8},
    'M200': {'pnl': 900, 'dit': 41, 'stability': 0.8},
    'SMSF': {'pnl': 600, 'dit': 40, 'stability': 0.8}
}

dynamic_benchmarks = load_strategy_config()
if not dynamic_benchmarks:
    dynamic_benchmarks = BASE_CONFIG.copy()

expired_df = pd.DataFrame()
if not df.empty and 'Status' in df.columns:
    expired_df = df[df['Status'] == 'Expired'].copy()

    if not expired_df.empty:
        hist_grp = expired_df.groupby('Strategy')
        for strat, grp in hist_grp:
            winners = grp[grp['P&L'] > 0]
            current_bench = dynamic_benchmarks.get(strat, {})
            if not winners.empty:
                current_bench['pnl'] = winners['P&L'].mean()
                current_bench['dit'] = winners['Days Held'].mean()
                current_bench['yield'] = grp['Daily Yield %'].mean()
                current_bench['roi'] = winners['ROI'].mean()
            dynamic_benchmarks[strat] = current_bench


def build_strategy_birdseye(all_df):
    if all_df.empty:
        return pd.DataFrame(columns=[
            'Strategy', 'Realized P&L', 'Floating P&L', 'Net P&L',
            'Delta', 'Theta', 'Active Trades', 'Closed Trades', 'Total Trades'
        ])

    active = all_df[all_df['Status'].isin(['Active', 'Missing'])].copy()
    closed = all_df[all_df['Status'] == 'Expired'].copy()

    active_agg = active.groupby('Strategy').agg({
        'P&L': 'sum',
        'Delta': 'sum',
        'Theta': 'sum',
        'id': 'count'
    }).rename(columns={
        'P&L': 'Floating P&L',
        'Delta': 'Delta',
        'Theta': 'Theta',
        'id': 'Active Trades'
    })

    closed_agg = closed.groupby('Strategy').agg({
        'P&L': 'sum',
        'id': 'count'
    }).rename(columns={
        'P&L': 'Realized P&L',
        'id': 'Closed Trades'
    })

    bird = active_agg.join(closed_agg, how='outer').fillna(0).reset_index()
    bird['Net P&L'] = bird['Realized P&L'] + bird['Floating P&L']
    bird['Total Trades'] = bird['Active Trades'] + bird['Closed Trades']

    for col in ['Realized P&L', 'Floating P&L', 'Net P&L', 'Delta', 'Theta']:
        bird[col] = pd.to_numeric(bird[col], errors='coerce').fillna(0.0)

    for col in ['Active Trades', 'Closed Trades', 'Total Trades']:
        bird[col] = pd.to_numeric(bird[col], errors='coerce').fillna(0).astype(int)

    return bird.sort_values('Net P&L', ascending=False)


def build_realized_equity_curve(expired_trades, initial_capital):
    if expired_trades.empty or initial_capital <= 0:
        return pd.DataFrame()

    curve = expired_trades.dropna(subset=['Exit Date']).copy()
    if curve.empty:
        return pd.DataFrame()

    curve['Exit Date'] = pd.to_datetime(curve['Exit Date']).dt.date
    day_pnl = curve.groupby('Exit Date')['P&L'].sum().sort_index().reset_index()
    day_pnl['Equity'] = initial_capital + day_pnl['P&L'].cumsum()
    day_pnl['Running Max'] = day_pnl['Equity'].cummax()
    day_pnl['Drawdown %'] = np.where(
        day_pnl['Running Max'] > 0,
        ((day_pnl['Equity'] - day_pnl['Running Max']) / day_pnl['Running Max']) * 100,
        0.0
    )
    return day_pnl


birdseye_df = build_strategy_birdseye(df)
strategy_list = birdseye_df['Strategy'].tolist()
if 'selected_strategy' not in st.session_state:
    st.session_state['selected_strategy'] = strategy_list[0] if strategy_list else None
elif strategy_list and st.session_state['selected_strategy'] not in strategy_list:
    st.session_state['selected_strategy'] = strategy_list[0]


tab_overview, tab_strategy, tab_analytics, tab_config = st.tabs([
    " Overview",
    " Strategy Drilldown",
    " Portfolio Analytics",
    " Strategy Config"
])

with tab_overview:
    if df.empty or 'Status' not in df.columns:
        st.info("Database is empty. Sync your workbooks to begin.")
    else:
        active_df = df[df['Status'].isin(['Active', 'Missing'])].copy()
        realized_pnl = float(expired_df['P&L'].sum()) if not expired_df.empty else 0.0
        floating_pnl = float(active_df['P&L'].sum()) if not active_df.empty else 0.0
        net_pnl = realized_pnl + floating_pnl

        total_delta = float(active_df['Delta'].sum()) if not active_df.empty else 0.0
        total_theta = float(active_df['Theta'].sum()) if not active_df.empty else 0.0
        var_95_dollars = calculate_portfolio_var(expired_df, total_cap)

        k1, k2, k3 = st.columns(3)
        k4, k5, k6 = st.columns(3)
        k1.metric("Realized P&L", f"${realized_pnl:,.0f}")
        k2.metric("Floating P&L", f"${floating_pnl:,.0f}")
        k3.metric("Total Net P&L", f"${net_pnl:,.0f}")
        k4.metric("Net Delta (Open)", f"{total_delta:,.2f}")
        k5.metric("Net Theta (Open)", f"{total_theta:,.2f}")
        k6.metric("VaR 95% (1D, Realized)", f"${var_95_dollars:,.0f}")

        c1, c2 = st.columns(2)

        with c1:
            st.markdown("##### P&L Mix by Strategy")
            if not birdseye_df.empty:
                mix_df = birdseye_df[['Strategy', 'Realized P&L', 'Floating P&L']].melt(
                    id_vars='Strategy',
                    value_vars=['Realized P&L', 'Floating P&L'],
                    var_name='Type',
                    value_name='P&L'
                )
                fig_mix = px.bar(
                    mix_df,
                    x='Strategy',
                    y='P&L',
                    color='Type',
                    barmode='group',
                    color_discrete_map={'Realized P&L': '#10b981', 'Floating P&L': '#38bdf8'},
                    title="Realized vs Floating"
                )
                fig_mix = apply_chart_theme(fig_mix)
                st.plotly_chart(fig_mix, use_container_width=True)
            else:
                st.info("No strategy data available yet.")

        with c2:
            st.markdown("##### Open Capital Allocation")
            if not active_df.empty:
                alloc_df = active_df.groupby('Strategy', as_index=False)['Debit'].sum()
                fig_alloc = px.pie(
                    alloc_df,
                    names='Strategy',
                    values='Debit',
                    hole=0.45,
                    title="Open Debit by Strategy"
                )
                fig_alloc = apply_chart_theme(fig_alloc)
                st.plotly_chart(fig_alloc, use_container_width=True)
            else:
                st.info("No active trades to visualize.")

        st.markdown("##### Strategy Bird's Eye")
        if not birdseye_df.empty:
            styled = birdseye_df.style.format({
                'Realized P&L': "${:,.0f}",
                'Floating P&L': "${:,.0f}",
                'Net P&L': "${:,.0f}",
                'Delta': "{:,.2f}",
                'Theta': "{:,.2f}"
            }).map(
                lambda v: 'color: #10b981; font-weight: 700' if isinstance(v, (int, float)) and v > 0 else (
                    'color: #f43f5e; font-weight: 700' if isinstance(v, (int, float)) and v < 0 else ''
                ),
                subset=['Realized P&L', 'Floating P&L', 'Net P&L']
            )
            st.dataframe(styled, hide_index=True, use_container_width=True)

            selected_idx = strategy_list.index(st.session_state['selected_strategy']) if st.session_state['selected_strategy'] in strategy_list else 0
            selected = st.radio(
                "Click a strategy to drill down",
                strategy_list,
                index=selected_idx,
                horizontal=True,
                key="selected_strategy_radio"
            )
            st.session_state['selected_strategy'] = selected

with tab_strategy:
    if df.empty or not strategy_list:
        st.info("No strategy data available.")
    else:
        selected_strategy = st.session_state.get('selected_strategy') or strategy_list[0]
        selected_idx = strategy_list.index(selected_strategy) if selected_strategy in strategy_list else 0
        selected_strategy = st.selectbox("Strategy", strategy_list, index=selected_idx, key="strategy_selector")
        st.session_state['selected_strategy'] = selected_strategy

        strat_df = df[df['Strategy'] == selected_strategy].copy().sort_values('Entry Date', ascending=False)
        strat_active = strat_df[strat_df['Status'].isin(['Active', 'Missing'])].copy()
        strat_closed = strat_df[strat_df['Status'] == 'Expired'].copy()

        s1, s2, s3, s4, s5 = st.columns(5)
        s1.metric("Trades", len(strat_df))
        s2.metric("Open", len(strat_active))
        s3.metric("Closed", len(strat_closed))
        s4.metric("Net P&L", f"${strat_df['P&L'].sum():,.0f}")
        s5.metric("Open Greeks", f"Δ {strat_active['Delta'].sum():,.2f} | Θ {strat_active['Theta'].sum():,.2f}")

        v1, v2 = st.columns(2)
        with v1:
            if not strat_df.empty:
                fig_scatter = px.scatter(
                    strat_df,
                    x='Days Held',
                    y='P&L',
                    color='Status',
                    size='Debit',
                    hover_data=['Name', 'Theta', 'Delta', 'Gamma', 'Vega'],
                    title=f"{selected_strategy}: Trade Distribution"
                )
                fig_scatter.add_hline(y=0, line_dash='dash', line_color='#94a3b8', opacity=0.6)
                fig_scatter = apply_chart_theme(fig_scatter)
                st.plotly_chart(fig_scatter, use_container_width=True)

        with v2:
            if not strat_closed.empty:
                eq = strat_closed.dropna(subset=['Exit Date']).sort_values('Exit Date').copy()
                eq['Cumulative Realized'] = eq['P&L'].cumsum()
                fig_eq = px.line(
                    eq,
                    x='Exit Date',
                    y='Cumulative Realized',
                    markers=True,
                    title=f"{selected_strategy}: Realized Equity"
                )
                fig_eq = apply_chart_theme(fig_eq)
                st.plotly_chart(fig_eq, use_container_width=True)
            else:
                st.info("No closed trades in this strategy yet.")

        st.markdown("##### Individual Trades (OptionStrat links preserved)")
        trade_cols = [
            'Name', 'Status', 'Entry Date', 'Exit Date', 'Days Held',
            'lot_size', 'P&L', 'Debit', 'ROI', 'Daily Yield %', 'Ann. ROI',
            'Theta', 'Delta', 'Gamma', 'Vega', 'IV',
            'Put P&L', 'Call P&L', 'Grade', 'Reason', 'Notes', 'Tags', 'Parent ID', 'Link'
        ]
        show_df = strat_df[trade_cols].copy()
        show_df['Entry Date'] = pd.to_datetime(show_df['Entry Date']).dt.date
        show_df['Exit Date'] = pd.to_datetime(show_df['Exit Date']).dt.date

        st.dataframe(
            show_df,
            hide_index=True,
            use_container_width=True,
            column_config={
                'P&L': st.column_config.NumberColumn('P&L', format='$%0.0f'),
                'Debit': st.column_config.NumberColumn('Debit', format='$%0.0f'),
                'ROI': st.column_config.NumberColumn('ROI', format='%.2f%%'),
                'Daily Yield %': st.column_config.NumberColumn('Daily Yield', format='%.2f%%'),
                'Ann. ROI': st.column_config.NumberColumn('Ann. ROI', format='%.1f%%'),
                'Theta': st.column_config.NumberColumn('Theta', format='%.2f'),
                'Delta': st.column_config.NumberColumn('Delta', format='%.2f'),
                'Gamma': st.column_config.NumberColumn('Gamma', format='%.3f'),
                'Vega': st.column_config.NumberColumn('Vega', format='%.2f'),
                'IV': st.column_config.NumberColumn('IV', format='%.3f'),
                'Link': st.column_config.LinkColumn('OptionStrat', display_text='Open')
            }
        )

with tab_analytics:
    if df.empty or expired_df.empty:
        st.info("Need more closed trades for portfolio analytics.")
    else:
        curve_df = build_realized_equity_curve(expired_df, total_cap)
        if curve_df.empty:
            st.info("No realized curve available yet.")
        else:
            a1, a2 = st.columns(2)
            with a1:
                st.metric("Max Drawdown", f"{curve_df['Drawdown %'].min():.2f}%")
            with a2:
                st.metric("Current Drawdown", f"{curve_df['Drawdown %'].iloc[-1]:.2f}%")

            fig_eq = px.line(curve_df, x='Exit Date', y='Equity', title='Realized Equity Curve')
            fig_eq = apply_chart_theme(fig_eq)
            st.plotly_chart(fig_eq, use_container_width=True)

            fig_dd = px.area(curve_df, x='Exit Date', y='Drawdown %', title='Realized Drawdown (%)')
            fig_dd.add_hline(y=0, line_dash='dash', line_color='#94a3b8', opacity=0.6)
            fig_dd = apply_chart_theme(fig_dd)
            st.plotly_chart(fig_dd, use_container_width=True)

        active_df = df[df['Status'].isin(['Active', 'Missing'])].copy()
        if not active_df.empty:
            c_left, c_right = st.columns(2)
            with c_left:
                open_pnl = active_df.groupby('Strategy', as_index=False)['P&L'].sum().sort_values('P&L', ascending=False)
                fig_open = px.bar(open_pnl, x='Strategy', y='P&L', color='P&L', color_continuous_scale='RdYlGn', title='Open P&L by Strategy')
                fig_open = apply_chart_theme(fig_open)
                st.plotly_chart(fig_open, use_container_width=True)
            with c_right:
                greek = pd.DataFrame({
                    'Greek': ['Delta', 'Theta', 'Gamma', 'Vega'],
                    'Value': [
                        active_df['Delta'].sum(),
                        active_df['Theta'].sum(),
                        active_df['Gamma'].sum(),
                        active_df['Vega'].sum()
                    ]
                })
                fig_greek = px.bar(greek, x='Greek', y='Value', color='Greek', title='Portfolio Greek Snapshot')
                fig_greek = apply_chart_theme(fig_greek)
                st.plotly_chart(fig_greek, use_container_width=True)

        snaps = load_snapshots()
        if not snaps.empty:
            st.markdown("##### Snapshot Trend")
            snap_daily = snaps.groupby('snapshot_date', as_index=False).agg({'pnl': 'sum', 'theta': 'sum'})
            fig_snap = px.line(snap_daily, x='snapshot_date', y=['pnl', 'theta'], title='Daily Snapshot Totals (PnL and Theta)')
            fig_snap = apply_chart_theme(fig_snap)
            st.plotly_chart(fig_snap, use_container_width=True)

with tab_config:
    st.markdown("### Strategy Configuration Manager")
    conn = get_db_connection()
    try:
        strat_df = pd.read_sql("SELECT * FROM strategy_config", conn)
        expected_cols = {
            'name': 'Name',
            'identifier': 'Identifier',
            'target_pnl': 'Target PnL',
            'target_days': 'Target Days',
            'min_stability': 'Min Stability',
            'description': 'Description',
            'typical_debit': 'Typical Debit'
        }
        for db_col in expected_cols.keys():
            if db_col not in strat_df.columns:
                strat_df[db_col] = 0.0 if 'pnl' in db_col or 'debit' in db_col else (0 if 'days' in db_col else "")
        strat_df = strat_df[list(expected_cols.keys())].rename(columns=expected_cols)

        edited_strats = st.data_editor(
            strat_df,
            num_rows="dynamic",
            key="strat_editor_main",
            use_container_width=True,
            column_config={
                "Name": st.column_config.TextColumn("Strategy Name", help="Unique name"),
                "Identifier": st.column_config.TextColumn("Keyword Match"),
                "Target PnL": st.column_config.NumberColumn("Profit Target ($)", format="$%d"),
                "Target Days": st.column_config.NumberColumn("Target DIT (Days)"),
                "Min Stability": st.column_config.NumberColumn("Min Stability", format="%.2f"),
                "Typical Debit": st.column_config.NumberColumn("Typical Debit ($)", format="$%d"),
                "Description": st.column_config.TextColumn("Notes")
            }
        )

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            if st.button("Save Changes"):
                if update_strategy_config(edited_strats):
                    st.success("Configuration Saved!")
                    st.cache_data.clear()
                    st.rerun()
        with c2:
            if st.button("Reprocess 'Other' Trades"):
                count = reprocess_other_trades()
                st.success(f"Reprocessed {count} trades!")
                st.cache_data.clear()
                st.rerun()
        with c3:
            if st.button("Reset to Defaults", type="secondary"):
                seed_default_strategies(force_reset=True)
                st.cache_data.clear()
                st.rerun()
    except Exception as e:
        st.error(f"Error loading strategies: {e}")
    finally:
        conn.close()

    st.caption("Allantis Trade Guardian v151.0")
